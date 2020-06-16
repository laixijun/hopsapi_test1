#encoding=utf-8
'''
Created on 2019年11月9日

@author: 瞌睡蟲子
'''
import xlwt
import xlrd
# from xlwt import BIFFRecords,Bitmap
from xlutils.compat import xrange
from xlutils.filter import XLWTWriter,XLRDReader,BaseFilter,BaseWriter,process
from openpyxl import utils
import os, hashlib

class Excel03:
    
    def __init__(self):
        self.__wt=None;
        self.__rd=None;
        self.__style_list=None;
        self.__path=None;

    # 创建xlsx文件
    def CreateExcel(self, sPath):
        self.__wt=xlwt.Workbook();
        self.__wt.add_sheet('Sheet1',cell_overwrite_ok=True);
        self.__wt.save(sPath);
        return self.OpenExcel(sPath);
        
    # 打开xlsx文件
    def OpenExcel(self, sPath):
        self.__rd = xlrd.open_workbook(sPath, formatting_info=True);
        w = XLWTWriter();
        process(XLRDReader(self.__rd, 'unknown.xls'), w);
        self.__wt = w.output[0][1];
        self.__style_list = w.style_list;
        self.__path=sPath;
        return self;
     
    # 保存xlsx文件
    def Save(self, bSave=True):
        if bSave:
            self.__wt.save(self.__path);
 
    # 关闭Excel 
    def CloseExcel(self, bSave=False):
        # 写副本保存，没有close
        self.Save(bSave);        
        # 读副本释放资源
        self.__rd.release_resources();
         
    # 创建工作表 
    def CreateSheet(self, strSheetName, strWhere=None, bSave=False):
        self.__wt.add_sheet(strSheetName, cell_overwrite_ok=True);
        self.Save(bSave);
        if strWhere and type(strWhere) == int:
            self.__handleSheet(strSheetName, mvIndex=strWhere, option=1);
            self.Save(bSave);    
         
    # 获取所有工作表名 
    def GetSheetsName(self):
        return self.__rd.sheet_names();        
          
    # 重命名工作表 
    def SheetRename(self, sheet, strNewName, bSave=False):
        self._getWtSheet(sheet).name = strNewName;
        self.Save(bSave);
         
    # 复制工作表 
    def CopySheet(self, sheet, strNewSheetName, bSave=False):
        self.__handleSheet(sheet, wtSheetname=strNewSheetName, option=2);
        self.Save(bSave);
              
    # 删除工作表 
    def DeleteSheet(self, sheet, bSave=False): 
        self.__handleSheet(sheet, option=3);
        self.Save(bSave);
         
    # 激活工作表 
    def ActiveSheet(self, sheet):
        tempSheet=self._getWtSheet(sheet);
        index=self.__wt.sheet_index(tempSheet.name);
        self.__wt.set_active_sheet(index);
         
    # 写入单元格 
    def WriteCell(self, sheet, strCell, data, bSave=False):  
        row, col, value, style = self._copyStyle(sheet, strCell);
        tempSheet=self._getWtSheet(sheet);
        tempSheet.write(row, col, data, style);
        self.Save(bSave);        
         
    # 读取单元格 
    def ReadCell(self, sheet, strCell): 
        tempSheet=self._getRdSheet(sheet);
        col, row=self._getCells(strCell);
        return tempSheet.cell_value(row, col);
         
    # 写入行 
    def WriteRow(self, sheet, strCell, data, bSave=False):        
        tempSheet=None;  
        col, row =self._getCells(strCell);
        isrefresh=True;
        for r in range(len(data)):
            row1, col1, value, style = self._copyStyle(sheet, [row+1, col+r+1],isrefresh);            
            # 必须在_copyStyle后面调用_getWtSheet，否则sheet对象已经被释放
            if isrefresh:
                tempSheet=self._getWtSheet(sheet);
                isrefresh=False;
            tempSheet.write(row1, col1, data[r], style);
        self.Save(bSave);
         
    # 写入列 
    def WriteColumn(self, sheet, strCell, data, bSave=False):
        tempSheet=None;  
        col, row =self._getCells(strCell);
        isrefresh=True;
        for r in range(len(data)):
            row1, col1, value, style = self._copyStyle(sheet, [row+r+1, col+1],isrefresh);     
            # 必须在_copyStyle后面调用_getWtSheet，否则sheet对象已经被释放
            if isrefresh:
                tempSheet=self._getWtSheet(sheet);
                isrefresh=False;
            tempSheet.write(row1, col1, data[r], style);
        self.Save(bSave);
         
    # 读取行 
    def ReadRow(self, sheet, strCell):
        col, row = self._getCells(strCell);
        return self._getRdSheet(sheet).row_values(row,col);
                
         
    # 读取列 
    def ReadColumn(self, sheet, strCell):
        col, row = self._getCells(strCell);
        return self._getRdSheet(sheet).col_values(col,row);
         
    # 插入行 
    def InsertRow(self, sheet, strCell, data, bSave=False):
        max_col=self.GetColumsCount(sheet);
        min_col, min_row = self._getCells(strCell);
        self.__handleCell(sheet, [min_row,0,min_row+1,max_col], way=False);
        self.WriteRow(sheet, strCell, data, False);
        self.Save(bSave);
         
    # 插入列 
    def InsertColumn(self, sheet, strCell, data, bSave=False):
        max_row=self.GetRowsCount(sheet);
        min_col,min_row = self._getCells(strCell);
        self.__handleCell(sheet, [0,min_col,max_row,min_col+1]);
        self.WriteColumn(sheet, strCell, data, bSave);
        self.Save(bSave);
     
    # 合并或拆分单元格 
    def MergeRange(self, sheet, strRange, option=True, bSave=False):
        tempSheet=self._getWtSheet(sheet);
        min_col, min_row, max_col, max_row=self._getCells(strRange);
        if option:
            tempSheet.merge(min_row, max_row-1, min_col, max_col-1);
        else:
            self.__handleMerge(sheet, [min_col, min_row, max_col, max_row]);
        self.Save(bSave);
 
    # 读取区域 
    def ReadRange(self, sheet, strRange):
        tempSheet=self._getRdSheet(sheet);        
        min_col, min_row, max_col, max_row=self._getCells(strRange);
        res=[]
        for row in range(min_row, max_row):
            res.append(tempSheet.row_values(row, min_col , max_col));
        return res;
             
    # 获取行数 
    def GetRowsCount(self, sheet):
        return self._getRdSheet(sheet).nrows;
      
    # 获取列数 
    def GetColumsCount(self, sheet):
        return self._getRdSheet(sheet).ncols;        
         
    # 删除行 
    def DeleteRow(self, sheet, strCell, bSave=False):
        max_col=self.GetColumsCount(sheet);
        min_col, min_row = self._getCells(strCell);
        self.__handleCell(sheet, [min_row,0,min_row+1,max_col], way=False, option=False);
        self.Save(bSave);                
         
    # 删除列 
    def DeleteColumn(self, sheet, strCell, bSave=False):
        max_row=self.GetRowsCount(sheet);
        min_col,min_row = self._getCells(strCell);
        self.__handleCell(sheet, [0,min_col,max_row,min_col+1], option=False);
        self.Save(bSave);        
         
    # 插入图片 ,只能bmp. fWidth, fHeight 表示缩放比例
    def InsertImage(self, sheet, strCell, sFilePath, fWidth, fHeight, bSave=False):
        col, row =self._getCells(strCell);
        tempSheet=self._getWtSheet(sheet);  
        tempSheet.insert_bitmap(sFilePath, row, col, 0, 0, fWidth, fHeight);
        self.Save(bSave);
           
#     # 删除图片 ，暂时没搞定，二进制数据段操作不过关。准备在读副本忘写副本copy的时候，从图片二进制数据中删除对应的数据段来实现
    def DeleteImage(self, sheet, objPic, bSave=False):
        raise Exception("xls暂未实现图片删除！")
#         col, row =self._getCells(strCell);
#         tempSheet = self._getWtSheet(sheet);
#         bmp = Bitmap.ImDataBmpRecord(sFilePath);
#         obj = Bitmap.ObjBmpRecord(row, col, tempSheet, bmp, 0, 0, fWidth, fHeight);
#         bmp=tempSheet.bmp_rec;
#         this_bmp = obj.get() + bmp.get();
#         if this_bmp in bmp:
#             print("11111")
#         else:
#             print("bbbb")
#         
#         self.Save(bSave);
          
    # 写入区域 
    def WriteRange(self, sheet, strCell, data, bSave=False):
        tempSheet=None;  
        col, row =self._getCells(strCell);
        isrefresh=True;
        for rg in range(len(data)):
            for r in range(len(data[rg])):
                row1, col1, value, style = self._copyStyle(sheet, [row+rg+1, col+r+1],isrefresh);
                # 必须在_copyStyle后面调用_getWtSheet，否则sheet对象已经被释放
                if not tempSheet:
                    tempSheet=self._getWtSheet(sheet);
                isrefresh=False;
                tempSheet.write(row1, col1, data[rg][r], style);
        self.Save(bSave);
         
    # 清除区域 
    def ClearRange(self, sheet, strRange, bClearFormat=True, bSave=False):
        min_col, min_row, max_col, max_row=self._getCells(strRange);
        isrefresh = True;
        tempSheet=self._getWtSheet(sheet);
        for row in range(min_row, max_row):
            for col in range(min_col, max_col):
                if bClearFormat:
                    tempSheet.write(row, col);
                else:
                    # 这里在批量写的时候，不要每读一次刷新一次副本。第一次读需要刷新读副本
                    row, col, value, style =self._copyStyle(sheet, [row, col], isrefresh);
                    # 必须在_copyStyle后面调用_getWtSheet，否则sheet对象已经被释放
                    if isrefresh:
                        tempSheet=self._getWtSheet(sheet);
                        isrefresh = False;
                    tempSheet.write(row, col, "", style);
        self.Save(bSave);
         
    # 删除区域 
    def DeleteRange(self, sheet, strRange, bSave=False):
        min_col, min_row, max_col, max_row=self._getCells(strRange);  
        self.__handleCell(sheet, [min_row, min_col, max_row, max_col], way=False, option=False);
        self.Save(bSave);
        
         
    # 设置行高 
    def SetRowHeight(self, sheet, strCell, fHeight, bSave=False):
        tempSheet=self._getWtSheet(sheet);
        col, row=self._getCells(strCell);
        hestyle = xlwt.easyxf('font:height '+ str((72 * fHeight)))
        tempSheet.row(row).set_style(hestyle);
        self.Save(bSave);   
 
    # 设置列宽 
    def SetColumnWidth(self, sheet, strCell, fWidth, bSave=False):
        tempSheet=self._getWtSheet(sheet);
        col, row=self._getCells(strCell); 
        tempSheet.col(col).width = int(256*72/84) * fWidth;
        self.Save(bSave);
         
    # 设置单元格颜色 
    def SetCellColor(self, sheet, strCell, listColor, bSave=False):
        row, col, value, style =self._copyStyle(sheet, strCell);
        style.pattern.pattern = xlwt.Pattern.SOLID_PATTERN;
        style.pattern.pattern_fore_colour = xlwt.Style.colour_map[self._getColor(listColor)];
        tempSheet=self._getWtSheet(sheet);
        tempSheet.write(row, col, value, style);
        self.Save(bSave);

    # 设置单元格字体颜色 
    def SetCellFontColor(self, sheet, strCell, listColor, bSave=False):
        row, col, value, style =self._copyStyle(sheet, strCell);
        style.font.colour_index = xlwt.Style.colour_map[self._getColor(listColor)];
        tempSheet=self._getWtSheet(sheet);
        tempSheet.write(row, col, value, style);
        self.Save(bSave);
         
         
    # 设置区域字体颜色 
    def SetRangeFontColor(self, sheet, strRange, listColor, bSave=False):
        min_col, min_row, max_col, max_row=self._getCells(strRange);
        for row in range(min_row, max_row):
            for col in range(min_col, max_col):
                self.SetCellFontColor(sheet, [row+1, col+1], listColor, bSave);

    # 设置区域颜色 
    def SetRangeColor(self, sheet, strRange, listColor, bSave=False):
        min_col, min_row, max_col, max_row=self._getCells(strRange);
        for row in range(min_row, max_row):
            for col in range(min_col, max_col):
                self.SetCellColor(sheet, [row+1, col+1], listColor, bSave);
    
    # 写副本     
    def _getWtSheet(self, sheet):
        return self.__wt.get_sheet(sheet);
    
    # 读副本
    def _getRdSheet(self, sheet, isrefresh=True):
        # 这里有个坑，由于读写分离是两个副本。因此每次写了之后要重新读取，才能读到上一个写的副本的内容。
        # 至于性能影响暂时没有测试,这里可能是个炸弹
        if isrefresh:
            self.CloseExcel(True);
            self.OpenExcel(self.__path);
        if type(sheet) == str:
            return self.__rd.sheet_by_name(sheet);
        elif type(sheet) == int:
            return self.__rd.sheet_by_index(sheet);
    
    # 表达式转坐标系 
    def _getCells(self, strCell):
        if type(strCell) == str:
            expTemp = strCell.upper();
            if ":" in expTemp:
                min_col, min_row, max_col, max_row = utils.cell.range_boundaries(expTemp);
                return (min_col-1, min_row-1, max_col, max_row);
            else:
                column,row=utils.cell.coordinate_from_string(expTemp);
                column = utils.column_index_from_string(column);
                return (column-1, row-1);
        elif type(strCell) == list and len(strCell) == 2:
            valueTemp = strCell[0];
            if type(valueTemp) == int:
                return (strCell[1]-1,strCell[0]-1);
            elif type(valueTemp) == list:
                return (strCell[0][1]-1, strCell[0][0]-1, strCell[1][1], strCell[1][0]);
            else:
                self._getException(strCell)
        else:
            self._getException(strCell);
    
    # 从xlrd复制单元格的样式到xlwt。这里增加一个是否刷新写副本的更新，当写了之后直接读的时候需要True，当批量写的时候为False，避免性能消耗（第一次读仍然需要True）。
    def _copyStyle(self, sheet, strCell, isrefresh=True):
        # 这里一个坑，必须先获取xlrd实例。保证读写同步，才能取样式。
        tempSheet = self._getRdSheet(sheet, isrefresh);
        col, row=self._getCells(strCell);
        max_col = tempSheet.ncols;
        max_row = tempSheet.nrows;        
        if col < max_col and row < max_row:
            style = self.__style_list[tempSheet.cell_xf_index(row,col)];
            value = tempSheet.cell_value(row, col);
            return row, col, value, style;
        else:
            return row, col, "", xlwt.Style.default_style;    

    # RGB添加
    def _getColor(self, color):
        colour_mark=None;
        if type(color) == str and len(color) == 6:
            colour_mark='COR_'+color.upper();
            color=[int('0x'+color[i:i+2].upper(),16) for i in range(0,len(color),2)];
        elif type(color) == list and len(color) == 3:
            tempList=list(filter(lambda x: type(x) == int, color));       
            if len(tempList)==3:
                colour_mark='COR_'+"".join([str(hex(r)).replace('0x','').upper() for r in color]);                 
            else:
                self._getException(color);
        else:
            self._getException(color);
        if colour_mark:
            if colour_mark not in xlwt.Style.colour_map.keys():
                for i in range(8,63):
                    if i not in xlwt.Style.colour_map.values():
                        xlwt.add_palette_colour(colour_mark, i);
                        self.__wt.set_colour_RGB(i, 251, 228, 228);
                        break;
            return colour_mark;
        else:
            self._getException(color);    
    
    # 从xlrd复制xls到xlwt，在复制过程中增删sheet 
    def __handleSheet(self, sheet, wtSheetname=None, mvIndex=-1, option=2):
        self.CloseExcel(True);
        self.__rd=xlrd.open_workbook(self.__path, formatting_info=True);
        w = HandleSheetFilter(rdSheetName=sheet, wtSheetname=wtSheetname, mvIndex=mvIndex, option=option);
        process(XLRDReader(self.__rd,'unknown.xls'), w);
        self.__wt=w.output[0][1];
        self.__style_list = w.style_list;
    
    # 从xlrd复制xls到xlwt，在复制过程中位移cell
    def __handleCell(self, sheet, cellRange, way=True, option=True):
        self.CloseExcel(True);
        self.__rd=xlrd.open_workbook(self.__path, formatting_info=True);
        w = HandleCellFilter(rdSheetName=sheet, cellRange=cellRange, way=way, option=option);
        process(XLRDReader(self.__rd,'unknown.xls'), w);
        self.__wt=w.output[0][1];
        self.__style_list = w.style_list;
        
    # 从xlrd复制xls到xlwt，在复制过程中位移cell
    def __handleMerge(self, sheet, cellRange):
        self.CloseExcel(True);
        self.__rd=xlrd.open_workbook(self.__path, formatting_info=True);
        w = HandleMergeFilter(rdSheetName=sheet, cellRange=cellRange);
        process(XLRDReader(self.__rd,'unknown.xls'), w);
        self.__wt=w.output[0][1];
        self.__style_list = w.style_list;
    
    # 参数错误
    def _getException(self,msg):
        raise Exception('参数错误：%s' % str(msg));

class HandleSheetFilter(BaseFilter,BaseWriter):
    """
    sheet表操作
    rdSheetName：要操作的表名
    wtSheetname：复制后的表名
    option：1：移动；2：复制，3：删除
    """

    def __init__(self,rdSheetName, wtSheetname=None, mvIndex=-1, option=2):
        self.__rdSheetName = rdSheetName;
        self.__wtSheetname = wtSheetname;
        self.__pading_sheet = None;
        self.__option = option;
        self.__mvIndex = mvIndex;
        self.__sheetIndex = -1;
        self.output = [];
    
    def sheetIndex(self,name):
        sheetNames=self.rdbook.sheet_names();
        for i in range(len(sheetNames)):
            if sheetNames[i] == name:
                return i;
        return -1;
        
    def workbook(self,rdbook,wtbook_name):
        self.rdbook = rdbook;
        if type(self.__rdSheetName) == int:
            self.__sheetIndex = self.__rdSheetName;
            self.__pading_sheet = self.rdbook.sheet_by_index(self.__rdSheetName);
            self.__rdSheetName = self.__pading_sheet.name;
        else:
            self.__pading_sheet = self.rdbook.sheet_by_name(self.__rdSheetName);
            self.__sheetIndex=self.sheetIndex(self.__rdSheetName);
        BaseWriter.workbook(self, rdbook, wtbook_name);
        
    def sheet(self,rdsheet,wtsheet_name):
        self.rdsheet = rdsheet;
        self.wtsheet_name = wtsheet_name;        
        if self.__sheetIndex > 0 and self.__mvIndex > 0:            
            myIndex=self.sheetIndex(rdsheet.name);
            if myIndex == self.__mvIndex:
                if self.__pading_sheet is None:
                    self.__pading_sheet = self.rdbook.sheet_by_name(self.__rdSheetName);
                BaseWriter.sheet(self, self.__pading_sheet, self.__rdSheetName);
                for row_x in xrange(self.__pading_sheet.nrows):
                    BaseWriter.row(self, row_x,row_x);
                    for col_x in xrange(self.__pading_sheet.row_len(row_x)):
                        BaseWriter.cell(self, row_x,col_x,row_x,col_x);       
                BaseWriter.sheet(self,rdsheet,wtsheet_name);
            elif self.__rdSheetName == rdsheet.name:
                self.__pading_sheet = rdsheet;
                # 复制sheet，并更改复制的位置
                if self.__option == 2:
                    BaseWriter.sheet(self, self.__pading_sheet, self.__wtSheetname);
            else:
                BaseWriter.sheet(self,rdsheet,wtsheet_name);        
        else:
            if self.__rdSheetName == rdsheet.name:
                self.__pading_sheet = rdsheet;
            # 删除sheet
            if self.__option < 3 or self.__rdSheetName != rdsheet.name:
                BaseWriter.sheet(self,rdsheet,wtsheet_name);
            # 复制sheet
            if self.__option == 2 and rdsheet.name == self.rdbook.sheet_by_index(-1).name:            
                BaseWriter.sheet(self, self.__pading_sheet, self.__wtSheetname);
                for row_x in xrange(self.__pading_sheet.nrows):
                    BaseWriter.row(self, row_x,row_x);
                    for col_x in xrange(self.__pading_sheet.row_len(row_x)):
                        BaseWriter.cell(self, row_x,col_x,row_x,col_x);        
        
    def close(self):
        if self.wtbook is not None:
            self.output.append((self.wtname,self.wtbook));
            del self.wtbook;


class HandleCellFilter(BaseFilter,BaseWriter):
    """
    cell 的位移操作
    rdSheetName：sheet表名
    cellRange：操作区域
    way: True:横向，False：纵向
    option：True：增，False：删
    """

    def __init__(self,rdSheetName,cellRange,way=True,option=True):
        self.__rdSheetName = rdSheetName;
        self.__n_cells = cellRange;
        self.__n_row = cellRange[2] - cellRange[0];
        self.__n_col = cellRange[3] - cellRange[1];
        self.__way = way;
        self.__option = option;
        self.output = [];
        
    def workbook(self,rdbook,wtbook_name):
        self.rdbook = rdbook;
        if type(self.__rdSheetName) == int:
            self.__rdSheetName = self.rdbook.sheet_by_index(self.__rdSheetName).name;
        BaseWriter.workbook(self, rdbook, wtbook_name);
        
    def sheet(self,rdsheet,wtsheet_name):
        self.rdsheet = rdsheet;
        BaseWriter.sheet(self,rdsheet,wtsheet_name);      
    
    def cell(self,rdrowx,rdcolx,wtrowx,wtcolx):
        if self.rdsheet.name == self.__rdSheetName:            
            # 插入cell
            if self.__option:
                # 横向增加区域
                if self.__way:   
                    # 操作的单元格在区间内
                    if rdcolx >= self.__n_cells[1] and self.__n_cells[0] <= rdrowx < self.__n_cells[2]:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx+self.__n_col);
                    else:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx);
                # 纵向增加区域
                else:             
                    # 操作的单元格在区间内
                    if rdrowx >= self.__n_cells[0] and self.__n_cells[1] <= rdcolx < self.__n_cells[3]:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx+self.__n_row,wtcolx);
                    else:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx);
            # 删除cell
            else:
                # 横向删减区域
                if self.__way:        
                    # 操作的单元格在区间内
                    if rdcolx >= self.__n_cells[3] and self.__n_cells[0] <= rdrowx < self.__n_cells[2]:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx-self.__n_col);
                    elif self.__n_cells[1] <= rdcolx < self.__n_cells[3] and self.__n_cells[0] <= rdrowx < self.__n_cells[2]:
                        pass;
                    else:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx);     
                # 纵向增加区域
                else:       
                    # 操作的单元格在区间内
                    if rdrowx >= self.__n_cells[2] and self.__n_cells[1] <= rdcolx < self.__n_cells[3]:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx-self.__n_row,wtcolx);
                        pass;
                    elif self.__n_cells[1]<= rdcolx < self.__n_cells[3] and self.__n_cells[0] <= rdrowx < self.__n_cells[2]:
                        pass;
                    else:
                        BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx);
        else:
            BaseWriter.cell(self,rdrowx,rdcolx,wtrowx,wtcolx);
        
        
    def close(self):
        if self.wtbook is not None:
            self.output.append((self.wtname,self.wtbook));
            del self.wtbook;

class HandleMergeFilter(BaseFilter,BaseWriter):
    """
    sheet 拆分单元格
    rdSheetName：要操作的表名
    cellRange：要拆分的单元格
    """

    def __init__(self,rdSheetName,cellRange):
        self.__rdSheetName = rdSheetName;
        self.__cellRange = cellRange;
        self.output = [];
        
    def workbook(self,rdbook,wtbook_name):
        self.rdbook = rdbook;
        if type(self.__rdSheetName) == int:
            self.__rdSheetName = self.rdbook.sheet_by_index(self.__rdSheetName).name;
        BaseWriter.workbook(self, rdbook, wtbook_name);
    
    def sheet(self,rdsheet,wtsheet_name):
        self.rdsheet = rdsheet;
        BaseWriter.sheet(self, rdsheet, wtsheet_name);
        if rdsheet.name == self.__rdSheetName:
            del self.merged_cell_top_left_map[(self.__cellRange[1],self.__cellRange[0])];

## 这里准备写取消合并单元格时，回填第一个单元格的值，暂时没搞定    
#     def cell(self,rdrowx,rdcolx,wtrowx,wtcolx):
#         BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx);
#         if rdrowx == self.__cellRange[0] and rdcolx == self.__cellRange[1]:
#             cell = self.rdsheet.cell(rdrowx,rdcolx);
#             style = xlwt.Style.default_style;
#             if cell.xf_index is not None:
#                 style = self.style_list[cell.xf_index]; 
#             print(cell.value);
#             print(style);
#             wtrow = self.wtsheet.row(rdrowx);           
#             wtrow.write(rdcolx, cell.value, style);
#         
    def close(self):
        if self.wtbook is not None:
            self.output.append((self.wtname,self.wtbook));
            del self.wtbook;   



# 定义UB调用，支持多个xls副本对象
APPS={};

# 创建xlsx文件
def CreateExcel(sPath):
    global APPS;
    m = hashlib.md5();
    m.update(os.path.normpath(sPath).encode("utf8"));
    key=m.hexdigest();
    if key not in APPS.keys():
        if os.path.isfile(sPath):
            return OpenExcel(sPath);
        else:
            stuff = sPath[-4:];
            stuff = stuff.upper();
            excel = None;
            if stuff == ".XLS":
                excel = Excel03().CreateExcel(sPath);
            else:
                raise Exception("文件格式错误！");
            APPS[key] = excel;
    return key;
        
# 打开xlsx文件
def OpenExcel(sPath):
    global APPS;
    m = hashlib.md5();
    m.update(os.path.normpath(sPath).encode("utf8"));
    key=m.hexdigest();
    if key not in APPS.keys():
        if os.path.isfile(sPath):
            stuff = sPath[-4:];
            stuff = stuff.upper();
            excel = None;
            if stuff == ".XLS":
                excel = Excel03().OpenExcel(sPath);
            else:
                raise Exception("文件格式错误！");
            APPS[key] = excel;
        else:
            return CreateExcel(sPath);
    return key;

# 保存xlsx文件
def Save(objExcelWorkBook):
    global APPS;
    APPS[objExcelWorkBook].Save();

# 关闭xlsx文件
def CloseExcel(objExcelWorkBook, bSave=True):
    global APPS;
    APPS[objExcelWorkBook].CloseExcel(bSave=bSave);
    del APPS[objExcelWorkBook];

# 创建工作表
def CreateSheet(objExcelWorkBook, strSheetName, strWhere=None, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].CreateSheet(strSheetName, strWhere=strWhere, bSave=bSave);

# 获取所有工作表名
def GetSheetsName(objExcelWorkBook):
    global APPS;
    return APPS[objExcelWorkBook].GetSheetsName();

# 重命名工作表
def SheetRename(objExcelWorkBook, sheet, strNewName, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SheetRename(sheet, strNewName, bSave=bSave);

# 复制工作表
def CopySheet(objExcelWorkBook, sheet, strNewSheetName, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].CopySheet(sheet, strNewSheetName, bSave=bSave);

# 删除工作表
def DeleteSheet(objExcelWorkBook, sheet, bSave=False):  
    global APPS;
    APPS[objExcelWorkBook].DeleteSheet(sheet, bSave=bSave);  

# 激活工作表
def ActiveSheet(objExcelWorkBook, sheet):
    global APPS;
    APPS[objExcelWorkBook].ActiveSheet(sheet);

# 写入单元格
def WriteCell(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteCell(sheet, strCell, data, bSave=bSave);

# 读取单元格
def ReadCell(objExcelWorkBook, sheet, strCell):
    global APPS;
    return APPS[objExcelWorkBook].ReadCell(sheet, strCell);

# 写入行
def WriteRow(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteRow(sheet, strCell, data, bSave=bSave);

# 写入列
def WriteColumn(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteColumn(sheet, strCell, data, bSave=bSave);

# 读取行
def ReadRow(objExcelWorkBook, sheet, strCell):
    global APPS;
    return APPS[objExcelWorkBook].ReadRow(sheet, strCell);

# 读取列
def ReadColumn(objExcelWorkBook, sheet, strCell):
    global APPS;
    return APPS[objExcelWorkBook].ReadColumn(sheet, strCell);

# 插入行
def InsertRow(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].InsertRow(sheet, strCell, data, bSave=bSave);

# 插入列
def InsertColumn(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].InsertColumn(sheet, strCell, data, bSave=bSave);

# 合并或拆分单元格
def MergeRange(objExcelWorkBook, sheet, strRange, option=True, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].MergeRange(sheet, strRange, option=option, bSave=bSave);

# 读取区域
def ReadRange(objExcelWorkBook, sheet, strRange):
    global APPS;
    return APPS[objExcelWorkBook].ReadRange(sheet, strRange);

# 获取行数
def GetRowsCount(objExcelWorkBook, sheet):
    global APPS;
    return APPS[objExcelWorkBook].GetRowsCount(sheet);

# 获取列数
def GetColumsCount(objExcelWorkBook, sheet):
    global APPS;
    return APPS[objExcelWorkBook].GetColumsCount(sheet);

# 删除行
def DeleteRow(objExcelWorkBook, sheet, strCell, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteRow(sheet, strCell, bSave=bSave);

# 删除列
def DeleteColumn(objExcelWorkBook, sheet, strCell, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteColumn(sheet, strCell, bSave=bSave);

# 插入图片
def InsertImage(objExcelWorkBook, sheet, strCell, sFilePath, fWidth, fHeight, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].InsertImage(sheet, strCell, sFilePath, fWidth, fHeight, bSave=bSave);

# 删除图片
def DeleteImage(objExcelWorkBook, sheet, objPic, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteImage(sheet, objPic, bSave=bSave);

# 写入区域
def WriteRange(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteRange(sheet, strCell, data, bSave=bSave);

# 清除区域
def ClearRange(objExcelWorkBook, sheet, strRange, bClearFormat=True, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].ClearRange(sheet, strRange, bClearFormat=bClearFormat, bSave=bSave);

# 删除区域
def DeleteRange(objExcelWorkBook, sheet, strRange, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteRange(sheet, strRange, bSave=bSave);

# 设置行高
def SetRowHeight(objExcelWorkBook, sheet, strCell, fHeight, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetRowHeight(sheet, strCell, fHeight, bSave=bSave);

# 设置列宽
def SetColumnWidth(objExcelWorkBook, sheet, strCell, fWidth, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetColumnWidth(sheet, strCell, fWidth, bSave=bSave);

# 设置单元格颜色
def SetCellColor(objExcelWorkBook, sheet, strCell, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetCellColor(sheet, strCell, listColor, bSave=bSave);

# 设置单元格字体颜色
def SetCellFontColor(objExcelWorkBook, sheet, strCell, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetCellFontColor(sheet, strCell, listColor, bSave=bSave);

# 设置区域字体颜色
def SetRangeFontColor(objExcelWorkBook, sheet, strRange, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetRangeFontColor(sheet, strRange, listColor, bSave=bSave);

# 设置区域颜色
def SetRangeColor(objExcelWorkBook, sheet, strRange, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetRangeColor(sheet, strRange, listColor, bSave=bSave);

			
if __name__ == '__main__':
    ""
# #     sPath=r"C:\Users\Administrator\Desktop\插件3.xlsm"
# #     print(str.lower(sPath[-4:]))
#     xls = Excel03();
# #     xls.CreateExcel(r"C:\Users\Administrator\Desktop\插件31.xls");
# #     xls.Save();
#     xls.OpenExcel(r"C:\Users\Administrator\Desktop\插件31.xls");
# #     xls.ActiveSheet("Sheet2");
# #     xls.CreateSheet("Sheet1",1,True);
# #     print(xls.GetSheetsName());
# #     xls.Save();
# #     xls.CloseExcel();
# #     xls.DeleteRow("Sheet5",[4,4],True);
# #     xls.SheetRename(0, "Sheet5",True);
# #     xls.CopySheet("Sheet5", "Sheet7", True);
# #     xls.DeleteSheet("Sheet2",True);
# #     xls.ActiveSheet("Sheet3");
# #     xls.WriteCell("Sheet1", "c8", "mmmm",True);
# #     data = xls.ReadCell(0, [8,3]);
# #     print(data);
# #     xls.WriteCell(0, [8,3], "2222",True);
# #     data = xls.ReadCell("Sheet1", "c8");
# #     print(data);
#     # print(data,type(data))
#     # xls.Save();
# #     xls.WriteRow("Sheet1", "b7", [111,"333","=SUM(A1:B21)"], True);
# #     xls.WriteColumn("Sheet1", "g8", [123,"3333A","=SUM(A1:B2)"], True);
# #     print(xls.ReadRow(0, "b7"))
# #     print(xls.ReadRow(0, "A5"))
# #     print(xls.ReadRow("Sheet1", [3,3]))
# #     print(xls.ReadColumn("Sheet1", [3,3]))
# #     print(xls.ReadColumn(0, "g8"))
# #     xls.InsertRow("Sheet1", [4,5], ["mmm","bbb",666,"rrr"], True);
# #     xls.InsertColumn(0, "C3", ["xxxxx","4444",666,"ddd"], True);
# #     print(xls.ReadColumn("Sheet1", [1,3]))
# #     print(xls.ReadColumn(0, "C2"))
# #     print(xls.ReadColumn(0, [3,4]))
# #     xls.MergeRange(0, "e3:i14", True, True);
# #     xls.MergeRange(0, "e3:i14", False, True);
# #     print(xls.ReadRange(0, "a3:F5"))
# #     print(xls.ReadRange(0, "c3:f5"))
# #     print(xls.ReadRange(0, "c3:f5"))
# #     print(xls.GetRowsCount("Sheet1"))
# #     print(xls.GetColumsCount(0))
# #     xls.DeleteRow("Sheet1", "D8", True)
# #     xls.DeleteColumn(0, "e7", True)
# #     xls.InsertImage(0, "d37", r"C:\Users\Administrator\Desktop\111.bmp", 0.5, 0.5,True)
# #     xls.InsertImage(0, "G26", r"C:\Users\Administrator\Desktop\222.bmp", 0.5, 0.5,True)
#     xls.DeleteImage(0, "G26", r"C:\Users\Administrator\Desktop\222.bmp", 0.5, 0.5, True);
# #     xls.WriteRange(0, "h47", [["111","1112","1113","1114","1115"],["111","1112","1113","1114","1115"],["111","1112","1113","1114","1115"],["111","1112","1113","1114","1115"]], True);
# #     xls.ClearRange("Sheet1", "c11:d14", True  , True);
# #     xls.DeleteRange(0, "d15:f18", True);
# #     xls.SetColumnWidth("Sheet1", "c8", 50 , True);
# #     xls.SetRowHeight("Sheet1", "c8", 50 , True);
# #     xls.SetCellColor("Sheet1", "c8", "AABBFF", True);
# #     xls.SetCellFontColor(0, "c8", [128,234,222], True);
# #     xls.SetRangeFontColor("Sheet1", "d3:e6", "AABBFF", True);
# #     xls.SetRangeColor(0, [[3,4],[6,5]], [128,234,222], True);
#      
#     xls.CloseExcel();
