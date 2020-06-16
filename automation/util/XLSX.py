#encoding=utf-8
'''
Created on 2019年11月9日

@author: 瞌睡蟲子
'''
import openpyxl
from openpyxl import utils,Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,PatternFill
import os,hashlib

class Excel07:
    def __init__(self):
        self.excel=None
        self.path=None

    # 创建xlsx文件
    def CreateExcel(self, sPath):
        self.excel=Workbook();
        self.excel["Sheet"].title = "Sheet1";
        self.excel.save(sPath);
        self.path=sPath;
        return self;
        
    # 打开xlsx文件
    def OpenExcel(self, sPath):
        stuff=str.lower(sPath[-4:]);
        if stuff=="xlsm":
            self.excel=openpyxl.load_workbook(sPath,keep_vba=True);
        else:
            self.excel=openpyxl.load_workbook(sPath);
        self.path=sPath;
        return self;
    
    # 保存xlsx文件
    def Save(self, bSave=True):
        if bSave:
            self.excel.save(self.path);
        
    # 创建工作表 
    def CreateSheet(self, strSheetName, strWhere=None, bSave=False):
        if strWhere:
            self.excel.create_sheet(strSheetName, strWhere);
        else:
            self.excel.create_sheet(strSheetName);
        self.Save(bSave);    
         
    # 获取所有工作表名 
    def GetSheetsName(self):
        return self.excel.sheetnames;        
         
    # 重命名工作表 
    def SheetRename(self, sheet, strNewName, bSave=False):
        self._getSheet(sheet).title = strNewName;
        self.Save(bSave);
        
    # 复制工作表 
    def CopySheet(self, sheet, strNewSheetName, bSave=False):
        tempSheet=self.excel.copy_worksheet(self._getSheet(sheet));
        tempSheet.title = strNewSheetName;
        self.Save(bSave);   
            
    # 删除工作表 
    def DeleteSheet(self, sheet, bSave=False):  
        self.excel.remove(self._getSheet(sheet));
        self.Save(bSave);
        
    # 激活工作表 
    def ActiveSheet(self, sheet):
        Workbook.active=self._getSheet(sheet);
        
    # 写入单元格 
    def WriteCell(self, sheet, strCell, data, bSave=False):
        cell=self._getCellOrRange(sheet, strCell);
        cell.value = data;
        self.Save(bSave);
        
    # 读取单元格 
    def ReadCell(self, sheet, strCell):
        cell=self._getCellOrRange(sheet, strCell);
        return cell.value;
        
    # 写入行 
    def WriteRow(self, sheet, strCell, data, bSave=False):
        tempSheet=self._getSheet(sheet);  
        tempCells=self._getCells(strCell);        
        column, row = utils.cell.coordinate_from_string(tempCells);
        column=utils.column_index_from_string(column);
        for k in range(len(data)):
            cell='%s%d'%(utils.get_column_letter(column+k),row);
            tempSheet[cell].value = data[k];
        self.Save(bSave);    
        
    # 写入列 
    def WriteColumn(self, sheet, strCell, data, bSave=False):
        tempSheet=self._getSheet(sheet);  
        tempCells=self._getCells(strCell);        
        column, row = utils.cell.coordinate_from_string(tempCells);
        for k in range(len(data)):
            cell='%s%d'%(column,row+k);
            tempSheet[cell].value = data[k];
        self.Save(bSave);
        
    # 读取行 
    def ReadRow(self, sheet, strCell):
        strRange=self._getRow(sheet, strCell);
        return self.ReadRange(sheet, strRange)[0];
        
    # 读取列 
    def ReadColumn(self, sheet, strCell):
        strRange=self._getColumn(sheet, strCell);
        tempRange=self.ReadRange(sheet, strRange)
        return [y for x in tempRange for y in x];
        
    # 插入行 
    def InsertRow(self, sheet, strCell, data, bSave=False):
        tempSheet=self._getSheet(sheet);
        if type(strCell)==int:
            tempSheet.insert_rows(strCell);
            strCell='%s%d'%("A",strCell);
        else:
            tempCells=self._getCells(strCell);        
            column, row = utils.cell.coordinate_from_string(tempCells);
            tempSheet.insert_rows(row);
            strCell=tempCells;
        self.WriteRow(sheet, strCell, data, bSave);  
        
    # 插入列 
    def InsertColumn(self, sheet, strCell, data, bSave=False):
        tempSheet=self._getSheet(sheet);
        if type(strCell)==int:
            tempSheet.insert_cols(strCell);
            strCell='%s%d'%(utils.get_column_letter(strCell),1);
        else:
            tempCells=self._getCells(strCell);        
            column, row = utils.cell.coordinate_from_string(tempCells);
            tempSheet.insert_cols(utils.column_index_from_string(column));
            strCell=tempCells;
        self.WriteColumn(sheet, strCell, data, bSave);
    
    # 合并或拆分单元格 
    def MergeRange(self, sheet, strRange, option=True, bSave=False):
        tempSheet=self._getSheet(sheet);
        tempCells=self._getCells(strRange);
        if option:
            tempSheet.merge_cells(tempCells);
        else:
            tempSheet.unmerge_cells(tempCells);        
        self.Save(bSave);

    # 关闭Excel 
    def CloseExcel(self, bSave=False):
        self.Save(bSave);        
        self.excel.close();

    # 读取区域 
    def ReadRange(self, sheet, strRange):
        cells = self._getCellOrRange(sheet, strRange);
        res=[]
        for r in cells:
            res.append([v.value for v in r]);
        return res;
            
    # 获取行数 
    def GetRowsCount(self, sheet):
        return self._getSheet(sheet).max_row;
    
    # 获取列数 
    def GetColumsCount(self, sheet):
        return self._getSheet(sheet).max_column;        
        
    # 删除行 
    def DeleteRow(self, sheet, strCell, bSave=False):
        tempSheet=self._getSheet(sheet);
        if type(strCell)==int:
            tempSheet.delete_rows(strCell);
        else:
            tempCells=self._getCells(strCell);        
            column, row = utils.cell.coordinate_from_string(tempCells);
            tempSheet.delete_rows(row);
        self.Save(bSave);  
        
    # 删除列 
    def DeleteColumn(self, sheet, strCell, bSave=False):
        tempSheet=self._getSheet(sheet);
        if type(strCell)==int:
            tempSheet.delete_cols(strCell);
        else:
            tempCells=self._getCells(strCell);        
            column, row = utils.cell.coordinate_from_string(tempCells);
            tempSheet.delete_cols(utils.column_index_from_string(column));
        self.Save(bSave);
        
    # 插入图片 
    def InsertImage(self, sheet, strCell, sFilePath, fWidth, fHeight, bSave=False):
        tempSheet=self._getSheet(sheet);  
        tempCells=self._getCells(strCell);        
        column, row = utils.cell.coordinate_from_string(tempCells);
        img = Image(sFilePath);
        img.width, img.height = (fWidth, fHeight);        
        tempSheet.column_dimensions[column].width = fWidth;
        tempSheet.row_dimensions[row].height = fHeight;
        tempSheet.add_image(img, tempCells);
        self.Save(bSave);
         
    # 删除图片 
    def DeleteImage(self, sheet, objPic, bSave=False):
        tempSheet = self._getSheet(sheet);
        del tempSheet._images[objPic];
        self.Save(bSave);
        
    # 写入区域 
    def WriteRange(self, sheet, strCell, data, bSave=False):
        tempSheet=self._getSheet(sheet);  
        tempCells=self._getCells(strCell);        
        column, row = utils.cell.coordinate_from_string(tempCells);
        column=utils.column_index_from_string(column);
        for rg in range(len(data)):
            for r in range(len(data[rg])):
                cell='%s%d'%(utils.get_column_letter(column+r),row+rg);
                tempSheet[cell].value = data[rg][r];
        self.Save(bSave);
        
    # 清除区域 
    def ClearRange(self, sheet, strRange, bClearFormat=True, bSave=False):
        cells=self._getCellOrRange(sheet, strRange);
        for row in cells:
            for cell in row:
                cell.value =None;
                if bClearFormat:
                    cell.font=Font();
                    cell.fill=PatternFill();
        self.Save(bSave);
        
    # 删除区域 
    def DeleteRange(self, sheet, strRange, bSave=False):
        tempRange=self._getCells(strRange);
        min_col, min_row, max_col, max_row1 = utils.cell.range_boundaries(tempRange);
        max_row=self.GetRowsCount(sheet);
        tempRange=self._getCells([[max_row1,min_col],[max_row,max_col]]);
        tempSheet=self._getSheet(sheet);
        tempSheet.move_range(tempRange, rows=(min_row-max_row1), cols=0, translate=True);
        self.ClearRange(sheet, [[max_row-(max_row1-min_row),min_col],[max_row,max_col]], True, False);        
        self.Save(bSave);
        
    # 设置行高 
    def SetRowHeight(self, sheet, strCell, fHeight, bSave=False):
        tempSheet=self._getSheet(sheet);  
        tempCells=self._getCells(strCell);        
        print(tempCells)
        column, row = utils.cell.coordinate_from_string(tempCells);
        tempSheet.row_dimensions[row].height = fHeight;
        self.Save(bSave);      

    # 设置列宽 
    def SetColumnWidth(self, sheet, strCell, fWidth, bSave=False):
        tempSheet=self._getSheet(sheet);  
        tempCells=self._getCells(strCell);        
        column, row = utils.cell.coordinate_from_string(tempCells);
        print(column)
        tempSheet.column_dimensions[column].width = fWidth;
        self.Save(bSave);
        
    # 设置单元格颜色 
    def SetCellColor(self, sheet, strCell, listColor, bSave=False):
        cell=self._getCellOrRange(sheet, strCell);
        cell.fill=PatternFill("solid", fgColor=self._getColor(listColor));
        self.Save(bSave);
        
    # 设置单元格字体颜色 
    def SetCellFontColor(self, sheet, strCell, listColor, bSave=False):
        cell=self._getCellOrRange(sheet, strCell);
        cell.font=Font(color=self._getColor(listColor));
        self.Save(bSave);
        
        
    # 设置区域字体颜色 
    def SetRangeFontColor(self, sheet, strRange, listColor, bSave=False):
        cells=self._getCellOrRange(sheet, strRange);
        tempColor=Font(color=self._getColor(listColor));
        for row in cells:
            for cell in row:
                cell.font=tempColor;
        self.Save(bSave);
        
    # 设置区域颜色 
    def SetRangeColor(self, sheet, strRange, listColor, bSave=False):
        cells=self._getCellOrRange(sheet, strRange);
        tempColor=PatternFill("solid", bgColor=self._getColor(listColor));
        for row in cells:
            for cell in row:
                cell.fill=tempColor;
        self.Save(bSave);        
        
    def _getSheet(self, sheet):
        if type(sheet) == str:
            return self.excel[sheet];
        elif type(sheet) == int:
            return self.excel.worksheets[sheet];
    
    def _getCells(self, strCell):
        if type(strCell) == str:
            return strCell.upper();
        elif type(strCell) == list and len(strCell) == 2:
            valueTemp = strCell[0];
            expTemp="";
            if type(valueTemp) == int:
                expTemp = '%s%d' % (utils.get_column_letter(strCell[1]), valueTemp);
            elif type(valueTemp) == list:
                expTemp = ['%s%d' % (utils.get_column_letter(r[1]), r[0]) for r in strCell]
                expTemp = ":".join(expTemp);
            else:
                self._getException(strCell)
            return expTemp.upper();
        else:
            self._getException(strCell)
    
    def _getColumn(self, sheet, strCell):
        tempSheet=self._getSheet(sheet);        
        if type(strCell) == str:
            column, row = utils.cell.coordinate_from_string(strCell);
            return '%s:%s%d' % (strCell,column,tempSheet.max_row);
        elif type(strCell) == list:
            tempList=list(filter(lambda x: type(x) == int, strCell));       
            if len(tempList)==2:
                strRange=[];
                strRange.append(strCell);
                strRange.append([tempSheet.max_row,strCell[1]]);
                return strRange;
            else:
                self._getException(strCell)
            
    def _getRow(self, sheet, strCell):
        tempSheet=self._getSheet(sheet);        
        if type(strCell) == str:
            column, row = utils.cell.coordinate_from_string(strCell);
            return '%s:%s%d' % (strCell,utils.get_column_letter(tempSheet.max_column),row);
        elif type(strCell) == list:
            tempList=list(filter(lambda x: type(x) == int, strCell));       
            if len(tempList)==2:
                strRange=[];
                strRange.append(strCell);
                strRange.append([strCell[0],tempSheet.max_column]);
                return strRange;
            else:
                self._getException(strCell)
    
    def _getCellOrRange(self, sheet, strCell):
        tempSheet=self._getSheet(sheet);
        tempCells=self._getCells(strCell);
        return tempSheet[tempCells];
    
    def _getColor(self, color):
        if type(color) == str and len(color) == 6:
            return color.upper();
        elif type(color) == list and len(color) == 3:
            tempList=list(filter(lambda x: type(x) == int, color));       
            if len(tempList)==3:
                return "".join([str(hex(r)).replace('0x','').upper() for r in color]); 
            else:
                self._getException(color);
        else:
            self._getException(color);
    
    def _getException(self,msg):
        raise Exception('参数错误：%s' % str(msg));



# 定义UB调用，支持多个xlsx,xlsm副本对象
APPS={};

# 创建xlsx文件
def CreateExcel(sPath):
    global APPS;
    m = hashlib.md5();
    m.update(os.path.normpath(sPath).encode("utf8"));
    key=m.hexdigest();
    if key not in APPS.keys():
        if os.path.isfile(sPath):
            return OpenExcel(sPath);
        else:
            stuff = sPath[-4:];
            stuff = stuff.upper();
            excel = None;
            if stuff == "XLSX" or stuff == "XLSM":
                excel = Excel07().CreateExcel(sPath);
            else:
                raise Exception("文件格式错误！");
            APPS[key] = excel;
    return key;
        
# 打开xlsx文件
def OpenExcel(sPath):
    global APPS;
    m = hashlib.md5();
    m.update(os.path.normpath(sPath).encode("utf8"));
    key=m.hexdigest();
    if key not in APPS.keys():
        if os.path.isfile(sPath):
            stuff = sPath[-4:];
            stuff = stuff.upper();
            excel = None;
            if stuff == "XLSX" or stuff == "XLSM":
                excel = Excel07().OpenExcel(sPath);
            else:
                raise Exception("文件格式错误！");
            APPS[key] = excel;
        else:
            return CreateExcel(sPath);
    return key;

# 保存xlsx文件
def Save(objExcelWorkBook):
    global APPS;
    APPS[objExcelWorkBook].Save();

# 关闭xlsx文件
def CloseExcel(objExcelWorkBook, bSave=True):
    global APPS;
    APPS[objExcelWorkBook].CloseExcel(bSave=bSave);
    del APPS[objExcelWorkBook];

# 创建工作表
def CreateSheet(objExcelWorkBook, strSheetName, strWhere=None, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].CreateSheet(strSheetName, strWhere=strWhere, bSave=bSave);

# 获取所有工作表名
def GetSheetsName(objExcelWorkBook):
    global APPS;
    return APPS[objExcelWorkBook].GetSheetsName();

# 重命名工作表
def SheetRename(objExcelWorkBook, sheet, strNewName, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SheetRename(sheet, strNewName, bSave=bSave);

# 复制工作表
def CopySheet(objExcelWorkBook, sheet, strNewSheetName, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].CopySheet(sheet, strNewSheetName, bSave=bSave);

# 删除工作表
def DeleteSheet(objExcelWorkBook, sheet, bSave=False):  
    global APPS;
    APPS[objExcelWorkBook].DeleteSheet(sheet, bSave=bSave);  

# 激活工作表
def ActiveSheet(objExcelWorkBook, sheet):
    global APPS;
    APPS[objExcelWorkBook].ActiveSheet(sheet);

# 写入单元格
def WriteCell(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteCell(sheet, strCell, data, bSave=bSave);

# 读取单元格
def ReadCell(objExcelWorkBook, sheet, strCell):
    global APPS;
    return APPS[objExcelWorkBook].ReadCell(sheet, strCell);

# 写入行
def WriteRow(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteRow(sheet, strCell, data, bSave=bSave);

# 写入列
def WriteColumn(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteColumn(sheet, strCell, data, bSave=bSave);

# 读取行
def ReadRow(objExcelWorkBook, sheet, strCell):
    global APPS;
    return APPS[objExcelWorkBook].ReadRow(sheet, strCell);

# 读取列
def ReadColumn(objExcelWorkBook, sheet, strCell):
    global APPS;
    return APPS[objExcelWorkBook].ReadColumn(sheet, strCell);

# 插入行
def InsertRow(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].InsertRow(sheet, strCell, data, bSave=bSave);

# 插入列
def InsertColumn(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].InsertColumn(sheet, strCell, data, bSave=bSave);

# 合并或拆分单元格
def MergeRange(objExcelWorkBook, sheet, strRange, option=True, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].MergeRange(sheet, strRange, option=option, bSave=bSave);

# 读取区域
def ReadRange(objExcelWorkBook, sheet, strRange):
    global APPS;
    return APPS[objExcelWorkBook].ReadRange(sheet, strRange);

# 获取行数
def GetRowsCount(objExcelWorkBook, sheet):
    global APPS;
    return APPS[objExcelWorkBook].GetRowsCount(sheet);

# 获取列数
def GetColumsCount(objExcelWorkBook, sheet):
    global APPS;
    return APPS[objExcelWorkBook].GetColumsCount(sheet);

# 删除行
def DeleteRow(objExcelWorkBook, sheet, strCell, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteRow(sheet, strCell, bSave=bSave);

# 删除列
def DeleteColumn(objExcelWorkBook, sheet, strCell, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteColumn(sheet, strCell, bSave=bSave);

# 插入图片
def InsertImage(objExcelWorkBook, sheet, strCell, sFilePath, fWidth, fHeight, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].InsertImage(sheet, strCell, sFilePath, fWidth, fHeight, bSave=bSave);

# 删除图片
def DeleteImage(objExcelWorkBook, sheet, objPic, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteImage(sheet, objPic, bSave=bSave);

# 写入区域
def WriteRange(objExcelWorkBook, sheet, strCell, data, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].WriteRange(sheet, strCell, data, bSave=bSave);

# 清除区域
def ClearRange(objExcelWorkBook, sheet, strRange, bClearFormat=True, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].ClearRange(sheet, strRange, bClearFormat=bClearFormat, bSave=bSave);

# 删除区域
def DeleteRange(objExcelWorkBook, sheet, strRange, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].DeleteRange(sheet, strRange, bSave=bSave);

# 设置行高
def SetRowHeight(objExcelWorkBook, sheet, strCell, fHeight, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetRowHeight(sheet, strCell, fHeight, bSave=bSave);

# 设置列宽
def SetColumnWidth(objExcelWorkBook, sheet, strCell, fWidth, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetColumnWidth(sheet, strCell, fWidth, bSave=bSave);

# 设置单元格颜色
def SetCellColor(objExcelWorkBook, sheet, strCell, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetCellColor(sheet, strCell, listColor, bSave=bSave);

# 设置单元格字体颜色
def SetCellFontColor(objExcelWorkBook, sheet, strCell, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetCellFontColor(sheet, strCell, listColor, bSave=bSave);

# 设置区域字体颜色
def SetRangeFontColor(objExcelWorkBook, sheet, strRange, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetRangeFontColor(sheet, strRange, listColor, bSave=bSave);

# 设置区域颜色
def SetRangeColor(objExcelWorkBook, sheet, strRange, listColor, bSave=False):
    global APPS;
    APPS[objExcelWorkBook].SetRangeColor(sheet, strRange, listColor, bSave=bSave);

                
if __name__ == '__main__':
    ""
    sPath=r"C:\Users\Administrator\Desktop\插件3.xlsm"
    print(str.lower(sPath[-4:]))
    xls = Excel07();
    xls.CreateExcel(r"C:\Users\Administrator\Desktop\插件3.xlsx");
    xls.Save();
    xls.CloseExcel();
    xls.OpenExcel(r"C:\Users\Administrator\Desktop\插件3.xlsx");
    print(xls.GetSheetsName());
    xls.CreateSheet("Sheet5",1,True);
    xls.Save();
    xls.DeleteRow("Sheet2",[1,1],True)
    xls.SheetRename(1, "strNewName1111",True);
    xls.CopySheet("strNewName", "hahah1", True);
    xls.DeleteSheet("hahah1");
    xls.ActiveSheet("Sheet3");
    xls.WriteCell(0, "B2", 111,True);
    data = xls.ReadCell(0, [2,2]);
    print(data,type(data))
    xls.Save();
    xls.WriteRow("strNewName", "A4", [123,"3333A","=SUM(A1:B2)"], True)
    xls.WriteColumn("strNewName", "A5", [123,"3333A","=SUM(A1:B2)"], True)
    print(xls.ReadRow(0, "A3"))
    print(xls.ReadRow(0, "a3"))
    print(xls.ReadRow(0, [3,1]))
    print(xls.ReadColumn("strNewName", [2,2]))
    xls.InsertRow(0, [4,5], ["mmm","bbb",666,"rrr"], True)
    xls.InsertColumn(0, "C3", ["xxxxx","4444",666,"ddd"], True)
    print(xls.ReadColumn("strNewName", [1,3]))
    print(xls.ReadColumn(0, "C2"))
    print(xls.ReadColumn(0, [3,4]))
    xls.MergeRange(0, "F14:h17", False, True);
    print(xls.ReadRange(0, [[3,3],[5,6]]))
    print(xls.ReadRange(0, "c3:f5"))
    print(xls.ReadRange(0, "c3:f5"))
    print(xls.GetRowsCount(0))
    print(xls.GetColumsCount("strNewName"))
    xls.DeleteRow(0, 35, True)
    xls.DeleteColumn(0, "E30", True)
    xls.InsertImage(0, "H21", r"C:\Users\Administrator\Desktop\11.jpg", 64, 64,True)
    xls.DeleteImage(0, 0, True);
    xls.WriteRange(0, "B24", [["111","1112","1113","1114","1115"],["111","1112","1113","1114","1115"],["111","1112","1113","1114","1115"],["111","1112","1113","1114","1115"]], True);
    xls.ClearRange("Sheet1", "B5:d8", True  , True);
    xls.DeleteRange(0, "b5:d8", True);
    xls.SetColumnWidth(0, [29,5], 0 , True);
    xls.SetRowHeight(0, [29,5], 0 , True);
    xls.SetCellColor(0, "A20", "1874CD", True)
    xls.SetCellFontColor(0, "A20", "FE44FF", True)
    xls.SetRangeFontColor("strNewName", "B22:e25", "FE44FF", True);
    xls.SetRangeColor("strNewName", "B24:e27", "1874cd", True);
    xls.CloseExcel();
