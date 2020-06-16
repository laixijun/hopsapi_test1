# _*_ coding: utf-8 _*_
# @Time         : 2019-03-17 14:26
# @Author        : 路培强
# @Email         : 136024009@qq.com
# @File          :ExcelUtil.py
# @Software      :PyCharm
from openpyxl import load_workbook


class DoExcel:
    """
    初始化数据，需要两个参数 处理Excel 采用openpyxl只能支持xlsx
    """
    def __init__(self, file_path, sheet_name):
        """
         file_path, sheet_name,group="FilePath"
        初始化数据，需要两个参数 处理Excel 采用openpyxl只能支持xlsx
        :param file_path:Excel文件路径
        :param sheet_name:Excel工作表单
        将配置文件提取出来，写活此文件
        """
        # conf = Config(group)
        # excelPath = conf.get_path_config("datas")
        # print(excelPath)
        # self.file_path = SystemOs().sys_path(excelPath, file_path)
        # self.file_path="E:/python_workspace/DestroyerRobot/automation/datas/test_data.xlsx"
        self.sheet_name = sheet_name
        self.file_path = file_path


    def do_excel(self, **dictData):
        """
        处理Excel 处理Excel 采用openpyxl只能支持xlsx
        :param dictData:需要传一个字典，{"变量1":列号,"变量2":列号}
        :return:
        """
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            for a, b in dictData.items():
                sub_data[a] = sheet.cell(i, b).value
                if sub_data[a] == None:
                    sub_data[a] = ""
            test_data.append(sub_data)
        return test_data
if __name__ == '__main__':
    # doExcel = DoExcel("test_data.xlsx","test_data")
    # #doExcel.do_excel({"1":"1"})
    #顺序从1开始计算
    dictData = {"userName": 3, "password": 4, "expected": 6}
    #test_data.xlsx 对应地址为 SystemOs().sys_path("automation/datas", file_path)
    #test_data 为脚表
    # filePath = "E:/python_workspace/DestroyerRobot/automation/com/cn/markerting_points/data/login/test_data_login.xlsx"
    # test_data = DoExcel(filePath, 'sheet1').do_excel(**dictData)
    # print(test_data)

    dictdata = {"goodsName": 3, "goodsImg": 7, "goodsInfo": 8, "goodsReject": 9}
    filepath = "E:\\python_workspace\\DestroyerRobot\\automation\\com\\cn\\markerting_points\\data\\goods\\goodsobject.xlsx"
    test = DoExcel(filepath,"sheet1").do_excel(**dictdata)
    print(test)


