# @Time    : 7/21/2020 6:19 PM
# @Author  : Yang Xiaobai
# @Email   : yangzhiyongtest@163.com
import decimal
import os
from decimal import Decimal
from time import sleep
import logging
import time
import pymysql

# from utils.config_tool.configurationEnv import DBSetting
# from utils.logger import Log
# from utils.new_tools.common_tool import Common
# from pyapi.scrapyApi.appium_app.get_driver import GetDriver
# from utils.spider_tool.time_wait import TimeWait
from appium import webdriver
import platform
import re
import openpyxl

# from utils.config_tool.ConfigParameter import WebSelenium
# from utils.config_tool.file_config_path import ExcelConfig
from selenium.common.exceptions import NoSuchElementException


class UiTest:
    def __init__(self):
        pass
    
    # #首页进入我的店铺的二维码并获取界面值，返回首页
    def testQRCodeStoreT(self, driver):
        sd = SpiderDemo()
        data = sd.testQRCodeStore(driver)
        return (data[1]["result"], data[1]["failDetail"])
    
    # 企业认证测试用例
    # ((<appium.webdriver.webdriver.WebDriver (session="d4d41c97-2a26-4979-85d2-1ed41957dd50")>, {'my': {'totle_money': '3,955,541.16', 'will_brokerage': '3,955,541.16', 'going_brokerage': '514,249.11', 'finish_brokerage': '1,036,347.92', 'cash': '63,780', 'card': '0'}, 'record': {}, 'withdraw': {'totle_money': '3,955,541.16', 'will_brokerage': '3,955,541.16', 'going_brokerage': '514,249.11', 'finish_brokerage': '1,036,347.92', 'cash': '63,780', 'card': '0'}, 'Outstanding': {'totle_money': '3,955,541.16', 'will_brokerage': '3,955,541.16', 'going_brokerage': '514,249.11', 'finish_brokerage': '1,036,347.92', 'cash': '63,780', 'card': '0'}, 'cash': {'totle_money': '3,955,541.16', 'will_brokerage': '3,955,541.16', 'going_brokerage': '514,249.11', 'finish_brokerage': '1,036,347.92', 'cash': '63,780', 'card': '0'}, 'shopping': {'totle_money': '3,955,541.16', 'will_brokerage': '3,955,541.16', 'going_brokerage': '514,249.11', 'finish_brokerage': '1,036,347.92', 'cash': '63,780', 'card': '0'}}), {'my': {'resultDetail': {'noPaidCommissionAmount': {'db': '3955541.16', 'ui': <class 'str'>}, 'paidingCommissionAmount': {'db': '514249.11', 'ui': <class 'str'>}, 'paidCommissionAmount': {'db': '1036347.92', 'ui': <class 'str'>}, 'noPaidCashAmount': {'db': '63780.00', 'ui': <class 'str'>}, 'noPaidShoppingCardAmount': {'db': '0.00', 'ui': <class 'str'>}}, 'noPaidCommissionAmount': 'fail', 'result': 'fail', 'paidingCommissionAmount': 'fail', 'paidCommissionAmount': 'fail', 'noPaidCashAmount': 'fail', 'noPaidShoppingCardAmount': 'fail'}, 'withdraw': {'resultDetail': {'noPaidCommissionAmount': {'db': '3955541.16', 'ui': <class 'str'>}, 'paidingCommissionAmount': {'db': '514249.11', 'ui': <class 'str'>}, 'paidCommissionAmount': {'db': '1036347.92', 'ui': <class 'str'>}, 'noPaidCashAmount': {'db': '63780.00', 'ui': <class 'str'>}, 'noPaidShoppingCardAmount': {'db': '0.00', 'ui': <class 'str'>}}, 'noPaidCommissionAmount': 'fail', 'result': 'fail', 'paidingCommissionAmount': 'fail', 'paidCommissionAmount': 'fail', 'noPaidCashAmount': 'fail', 'noPaidShoppingCardAmount': 'fail'}, 'Outstanding': {'resultDetail': {'noPaidCommissionAmount': {'db': '3955541.16', 'ui': <class 'str'>}, 'paidingCommissionAmount': {'db': '514249.11', 'ui': <class 'str'>}, 'paidCommissionAmount': {'db': '1036347.92', 'ui': <class 'str'>}, 'noPaidCashAmount': {'db': '63780.00', 'ui': <class 'str'>}, 'noPaidShoppingCardAmount': {'db': '0.00', 'ui': <class 'str'>}}, 'noPaidCommissionAmount': 'fail', 'result': 'fail', 'paidingCommissionAmount': 'fail', 'paidCommissionAmount': 'fail', 'noPaidCashAmount': 'fail', 'noPaidShoppingCardAmount': 'fail'}, 'cash': {'resultDetail': {'noPaidCommissionAmount': {'db': '3955541.16', 'ui': <class 'str'>}, 'paidingCommissionAmount': {'db': '514249.11', 'ui': <class 'str'>}, 'paidCommissionAmount': {'db': '1036347.92', 'ui': <class 'str'>}, 'noPaidCashAmount': {'db': '63780.00', 'ui': <class 'str'>}, 'noPaidShoppingCardAmount': {'db': '0.00', 'ui': <class 'str'>}}, 'noPaidCommissionAmount': 'fail', 'result': 'fail', 'paidingCommissionAmount': 'fail', 'paidCommissionAmount': 'fail', 'noPaidCashAmount': 'fail', 'noPaidShoppingCardAmount': 'fail'}, 'shopping': {'resultDetail': {'noPaidCommissionAmount': {'db': '3955541.16', 'ui': <class 'str'>}, 'paidingCommissionAmount': {'db': '514249.11', 'ui': <class 'str'>}, 'paidCommissionAmount': {'db': '1036347.92', 'ui': <class 'str'>}, 'noPaidCashAmount': {'db': '63780.00', 'ui': <class 'str'>}, 'noPaidShoppingCardAmount': {'db': '0.00', 'ui': <class 'str'>}}, 'noPaidCommissionAmount': 'fail', 'result': 'fail', 'paidingCommissionAmount': 'fail', 'paidCommissionAmount': 'fail', 'noPaidCashAmount': 'fail', 'noPaidShoppingCardAmount': 'fail'}})
    def testCompareEnterpriseCertificationT(self, driver, xinTax, phone):
        sd = SpiderDemo()
        # try:
        data = sd.compareEnterpriseCertification(driver, xinTax, phone)
        logger.info(data)
        return (data[1]["result"], data[1]["failDetail"])
    
    # 牵头人佣金测试用例
    # (<appium.webdriver.webdriver.WebDriver (session="eff97e03-605b-4f0f-8fbe-7769ae2c9c1d")>, {'failDetail': {'surplusAmount': {'expect': Decimal('17957517.54'), 'actual': Decimal('2085404.06')}}, 'shouldAmount': {'result': 'Success', 'detail': {'expect': Decimal('7936056.74'), 'actual': Decimal('7936056.74')}}, 'notAmount': {'result': 'Success', 'detail': {'expect': Decimal('7896056.74'), 'actual': Decimal('7896056.74')}}, 'auditAmount': {'result': 'Success', 'detail': {'expect': Decimal('40000.00'), 'actual': Decimal('40000.00')}}, 'alreadyAmount': {'result': 'Success', 'detail': {'expect': Decimal('0.00'), 'actual': Decimal('0.00')}}, 'estimateAmount': {'result': 'Success', 'detail': {'expect': Decimal('10021460.80'), 'actual': Decimal('10021460.80')}}, 'surplusAmount': {'result': 'fail', 'detail': {'expect': Decimal('17957517.54'), 'actual': Decimal('2085404.06')}}, 'result': 'Fail'})
    def testInitiatorCommissionAssert(self, driver):
        data = SpiderDemo().initiatorCommissionAssert(driver)
        data = data[1]
        myCommissionDic = (data['result'], data['failDetail'])
        return myCommissionDic
    
    # 我的佣金测试用例
    # ((<appium.webdriver.webdriver.WebDriver (session="c894af69-58ab-4e41-9743-41474b299c98")>, {'my': {'totle_money': '4,019,368.61', 'will_brokerage': '4,019,368.61', 'going_brokerage': '1,460,439.92', 'finish_brokerage': '1,046,286.96', 'cash': '13,370.00', 'card': '20,000.00'}, 'record': {}, 'withdraw': {'totle_money': '4,019,368.61', 'will_brokerage': '4,019,368.61', 'going_brokerage': '1,460,439.92', 'finish_brokerage': '1,046,286.96', 'cash': '13,370.00', 'card': '20,000.00'}, 'Outstanding': {'totle_money': '4,019,368.61', 'will_brokerage': '4,019,368.61', 'going_brokerage': '1,460,439.92', 'finish_brokerage': '1,046,286.96', 'cash': '13,370.00', 'card': '20,000.00'}, 'cash': {'totle_money': '4,019,368.61', 'will_brokerage': '4,019,368.61', 'going_brokerage': '1,460,439.92', 'finish_brokerage': '1,046,286.96', 'cash': '13,370.00', 'card': '20,000.00'}, 'shopping': {'totle_money': '4,019,368.61', 'will_brokerage': '4,019,368.61', 'going_brokerage': '1,460,439.92', 'finish_brokerage': '1,046,286.96', 'cash': '13,370.00', 'card': '20,000.00'}}), {'my': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'withdraw': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'Outstanding': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'cash': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'shopping': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}})
    # ['my','record','withdraw','Outstanding','cash','shopping']
    # {'my': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'withdraw': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'Outstanding': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'cash': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}, 'shopping': {'resultDetail': {'paidCommissionAmount': {'db': '1039276.96', 'ui': '1046286.96'}}, 'noPaidCommissionAmount': 'success', 'paidingCommissionAmount': 'success', 'paidCommissionAmount': 'fail', 'result': 'fail', 'noPaidCashAmount': 'success', 'noPaidShoppingCardAmount': 'success'}}
    def testMyCommissionAssert(self, driver):
        data = SpiderDemo().myCommissionAssert(driver)
        data = data[1]
        myCommissionDic = {}
        myCommissionDic['detail'] = {}
        for k, v in data.items():
            if v['result'] == 'fail':
                myCommissionDic['result'] = 'fail'
                myCommissionDic['detail'][k] = v['resultDetail']
            else:
                v['result'] = 'success'
        myCommissionDicTouple = (myCommissionDic['result'], myCommissionDic['detail'])
        return myCommissionDicTouple
class Log(object):
    """编写日志类，供其他模块调用"""
    
    def __init__(self, logger):
        # 拼接日志文件夹，如果不存在则自动创建
        report_path = str(os.path.dirname(__file__).split('utils')[0])
        # D:/zhy/haoshenghuo/autoproject/hopsapi_test1/pyapi/scrapyApi/appium_app\\report/logs
        log_path = os.path.join(report_path, 'logs')
        log_path = log_path.replace("\\", "/")
        if not os.path.exists(log_path):
            os.mkdir(log_path)
        
        # 设置日志文件名称格式及日志等级
        self.log_name = os.path.join(log_path, '%s.log' % time.strftime('%Y-%m-%d'))
        self.logger = logging.getLogger(logger)
        self.logger.setLevel(logging.DEBUG)
        
        # 创建一个handler，将日志写入日志文件
        fh = logging.FileHandler(self.log_name, 'a', encoding='utf-8')
        fh.setLevel(logging.INFO)
        
        # 再创建一个handler，将日志输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        
        # 定义日志输出格式
        self.formatter = logging.Formatter('%(asctime)s %(filename)s [line:%(lineno)d] %(levelname)s %(message)s')
        ch.setFormatter(self.formatter)
        fh.setFormatter(self.formatter)
        
        # 给logger添加handler
        self.logger.addHandler(fh)
        self.logger.addHandler(ch)
    
    def get_log(self):
        return self.logger
logger = Log(logger='deal_source').get_log()
class SpiderDemo:
    def __init__(self):
        # self.tm = TimeWait()
        pass
    
    # 启动APP，在首页
    def startApp(self, driver):
        driver.implicitly_wait(10)
        el1 = driver.find_element_by_id("com.android.packageinstaller:id/permission_allow_button")
        el1.click()
        el2 = driver.find_element_by_id("com.android.packageinstaller:id/permission_allow_button")
        el2.click()
        sleep(2)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvAgreement")
        el3.click()
        return driver
    
    # 登录APP,登录后在我的界面
    def loginApp(self, driver, phone="15517655129"):
        # driver = self.driver
        el4 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabMine")
        el4.click()
        elementExist = ToolsAppium().is_element_exist(driver=driver, element='常驻城市')
        if elementExist:
            el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/sure")
            el1.click()
        el5 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvlogin")
        el5.click()
        sleep(5)
        el6 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edphone")
        el6.send_keys(phone)
        el7 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btnLogin")
        el7.click()
        sleep(5)
        elementExist = ToolsAppium().is_element_exist(driver=driver, element='常驻城市')
        if elementExist:
            el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/sure")
            el1.click()
        sleep(3)
        return driver
    
    # 我的回到首页
    def mineGoHome(self, driver):
        el8 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabIndex")
        el8.click()
        logger.info("回到首页")
        return driver
    
    # 首页回到我的界面
    def homeGoMine(self, driver):
        sleep(2)
        el9 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabMine")
        el9.click()
        return driver
    
    # 首页进入我的店铺的二维码并获取界面值，返回首页
    def myStoreQR(self, driver):
        ta = ToolsAppium()
        flag = ta.isElement("id", ParamIdApp.USEAPP + ":id/tvIndexStoreCode", driver)
        if flag == False:
            logger.info(flag)
            driver = self.mineGoHome(driver)
        QRCodeDic = {}
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgQRCode")
        el1.click()
        sleep(3)
        # 获取用户名
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBrokerName").text
        QRCodeDic["broker_name"] = text
        # 获取手机号
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBrokerPhone").text
        QRCodeDic["broker_phone"] = text
        source = driver.page_source
        logger.info(source)
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgQRCode").text
        print(text)
        # 判断是否有二维码生成
        ta = ToolsAppium()
        flag = ta.isElement("id", ParamIdApp.USEAPP + ":id/imgQRCode", driver)
        QRCodeDic["imgQRCode"] = flag
        logger.info(flag)
        # 判断微信好友的元素出现
        flag = ta.isElement("id", ParamIdApp.USEAPP + ":id/tvShareWechat", driver)
        QRCodeDic["ShareWechat"] = flag
        logger.info(flag)
        # 判断朋友圈
        flag = ta.isElement("id", ParamIdApp.USEAPP + ":id/tvShareCircle", driver)
        QRCodeDic["ShareCircle"] = flag
        logger.info(flag)
        # 判断保存图片
        flag = ta.isElement("id", ParamIdApp.USEAPP + ":id/tvShareSave", driver)
        QRCodeDic["ShareSave"] = flag
        logger.info(flag)
        # 返回首页
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgFinish")
        el2.click()
        sleep(3)
        QRCodeDicDBKey = "broker_name" + "=" + QRCodeDic["broker_name"] + "!_" + "broker_phone" + "=" + QRCodeDic[
            "broker_phone"]
        QRCodeDicDBValue = "broker_phone" + "=" + QRCodeDic[
            "broker_phone"] + "!_" + "common" + "=" + "!0" + "!_" + "tableName" + "=" + "broker"
        QRCodeDicDB = {}
        QRCodeDicDB[QRCodeDicDBKey] = QRCodeDicDBValue
        return (driver, QRCodeDic, QRCodeDicDB)
    
    # 对比我的店铺二维码数据库值
    def compareQRCodeStore(self, QRCData):
        ms = ToolsAppium()
        for k, v in QRCData.items():
            result = ms.getDataExcept("t", k, v)
        return result
    
    # 店铺二维码测试执行方法
    def testQRCodeStore(self, driver):
        QRCodeDicList = ["imgQRCode", "ShareWechat", "ShareCircle", "ShareSave"]
        QRCodeResult = {}
        QRCodeResult["failDetail"] = {}
        sourceData = self.myStoreQR(driver)
        comResult = self.compareQRCodeStore(QRCData=sourceData[2])
        QRCodeDic = sourceData[1]
        for k, v in QRCodeDic.items():
            if k in QRCodeDicList:
                logger.info(v)
                if v == True:
                    QRCodeResult["resultList"][k] = "Success"
                else:
                    QRCodeResult["resultList"][k] = "Fail"
        if comResult["FAIL"] != {}:
            for k, v in comResult["FAIL"].items():
                QRCodeResult["failDetail"][k] = {"result": "Fail", "expect": comResult["expectDic"][k],
                                                 "actual": comResult["actule"][k]}
                QRCodeResult["result"] = "fail"
        else:
            QRCodeResult["result"] = "Success"
        
        logger.info(QRCodeResult)
        return (driver, QRCodeResult)
    
    # 我的钱包、关键数据立即提现、我的界面的待结佣、现金奖励、购物卡奖励入口
    # 我的佣金的校验
    def myCommission(self, driver):
        # 数据获取
        myCommissionDic = {}
        # 我的钱包
        myCommissionDic["my"] = {}
        # 结佣记录
        myCommissionDic["record"] = {}
        # 提现
        myCommissionDic["withdraw"] = {}
        # 待结佣
        myCommissionDic["Outstanding"] = {}
        # 现金奖励
        myCommissionDic["cash"] = {}
        # 购物卡奖励
        myCommissionDic["shopping"] = {}
        # 判断是否在首页，没有在首页的需要跳转到首页
        ta = ToolsAppium()
        # flag=ta.is_element_exist(driver,"待结佣")
        # if flag == False:
        #     #我的界面进入首页
        driver = self.mineGoHome(driver)
        # 我的钱包进入
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvIndexVolley")
        el1.click()
        sleep(3)
        # 待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_totle_money").text
        myCommissionDic["my"]["totle_money"] = text
        # 历史情况待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_will_brokerage").text
        myCommissionDic["my"]["will_brokerage"] = text
        # 历史情况申请中金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_going_brokerage").text
        myCommissionDic["my"]["going_brokerage"] = text
        # 历史情况已结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_finish_brokerage").text
        myCommissionDic["my"]["finish_brokerage"] = text
        # 经纪人奖励现金奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_cash").text
        myCommissionDic["my"]["cash"] = text
        # 经纪人奖励购物卡奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_card").text
        myCommissionDic["my"]["card"] = text
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/title")
        el2.click()
        # 结佣记录
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_brokerage_record")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        # 我的钱包回到首页
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/back")
        el3.click()
        sleep(3)
        driver = ToolsAppium().swipeUp01(driver)
        # 关键输入立即提现进入
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvRedraw")
        el1.click()
        sleep(5)
        # 待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_totle_money").text
        myCommissionDic["withdraw"]["totle_money"] = text
        # 历史情况待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_will_brokerage").text
        myCommissionDic["withdraw"]["will_brokerage"] = text
        # 历史情况申请中金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_going_brokerage").text
        myCommissionDic["withdraw"]["going_brokerage"] = text
        # 历史情况已结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_finish_brokerage").text
        myCommissionDic["withdraw"]["finish_brokerage"] = text
        # 经纪人奖励现金奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_cash").text
        myCommissionDic["withdraw"]["cash"] = text
        # 经纪人奖励购物卡奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_card").text
        myCommissionDic["withdraw"]["card"] = text
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_brokerage_record")
        el1.click()
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        # 关键输入立即提现回到首页
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/back")
        el2.click()
        sleep(3)
        # 进入我的界面
        driver = self.homeGoMine(driver)
        sleep(2)
        # 我的界面的待结佣进入
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/llWite")
        el1.click()
        sleep(5)
        # 待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_totle_money").text
        myCommissionDic["Outstanding"]["totle_money"] = text
        # 历史情况待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_will_brokerage").text
        myCommissionDic["Outstanding"]["will_brokerage"] = text
        # 历史情况申请中金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_going_brokerage").text
        myCommissionDic["Outstanding"]["going_brokerage"] = text
        # 历史情况已结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_finish_brokerage").text
        myCommissionDic["Outstanding"]["finish_brokerage"] = text
        # 经纪人奖励现金奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_cash").text
        myCommissionDic["Outstanding"]["cash"] = text
        # 经纪人奖励购物卡奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_card").text
        myCommissionDic["Outstanding"]["card"] = text
        # 结佣记录
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_brokerage_record")
        el1.click()
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        # 我的界面待结佣退出
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/back")
        el2.click()
        sleep(5)
        # 我的界面现金奖励进入
        el1 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/androidx.viewpager.widget.ViewPager/android.widget.LinearLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.TextView[2]")
        el1.click()
        sleep(3)
        # 待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_totle_money").text
        myCommissionDic["cash"]["totle_money"] = text
        # 历史情况待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_will_brokerage").text
        myCommissionDic["cash"]["will_brokerage"] = text
        # 历史情况申请中金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_going_brokerage").text
        myCommissionDic["cash"]["going_brokerage"] = text
        # 历史情况已结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_finish_brokerage").text
        myCommissionDic["cash"]["finish_brokerage"] = text
        # 经纪人奖励现金奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_cash").text
        myCommissionDic["cash"]["cash"] = text
        # 经纪人奖励购物卡奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_card").text
        myCommissionDic["cash"]["card"] = text
        # 结佣记录
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_brokerage_record")
        el1.click()
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        # 我的界面现金奖励出来
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/back")
        el2.click()
        sleep(3)
        
        # 我的购物卡奖励进入
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/llCard")
        el1.click()
        sleep(5)  # 待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_totle_money").text
        myCommissionDic["shopping"]["totle_money"] = text
        # 历史情况待结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_will_brokerage").text
        myCommissionDic["shopping"]["will_brokerage"] = text
        # 历史情况申请中金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_going_brokerage").text
        myCommissionDic["shopping"]["going_brokerage"] = text
        # 历史情况已结佣金额
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_finish_brokerage").text
        myCommissionDic["shopping"]["finish_brokerage"] = text
        # 经纪人奖励现金奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_cash").text
        myCommissionDic["shopping"]["cash"] = text
        # 经纪人奖励购物卡奖励
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_card").text
        myCommissionDic["shopping"]["card"] = text
        # 结佣记录
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_brokerage_record")
        el1.click()
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        # 我的购物卡奖励退出
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/back")
        el2.click()
        sleep(5)
        # 我的界面回到首页
        driver = self.mineGoHome(driver)
        return (driver, myCommissionDic)
    
    # 2020-11-19 17:12:40,128 broker_spider.py [line:1725] INFO {'noPaidCommissionAmount': Decimal('3955330.27'), '销售佣金': Decimal('917234.23'), '合伙佣金': Decimal('120996.60'), 'paidingCommissionAmount': Decimal('513960.00'), 'paidCommissionAmount': Decimal('1036347.92'), 'noPaidCashAmount': Decimal('63780.00'), 'noPaidShoppingCardAmount': Decimal('0.00')}
    # (<appium.webdriver.webdriver.WebDriver (session="54a4078a-faee-4c24-b9a4-b85b8b1ede37")>, {'my': {'totle_money': '0', 'will_brokerage': '0', 'going_brokerage': '0', 'finish_brokerage': '0', 'cash': '0', 'card': '0'}, 'record': {}, 'withdraw': {'totle_money': '0', 'will_brokerage': '0', 'going_brokerage': '0', 'finish_brokerage': '0', 'cash': '0', 'card': '0'}, 'Outstanding': {'totle_money': '0', 'will_brokerage': '0', 'going_brokerage': '0', 'finish_brokerage': '0', 'cash': '0', 'card': '0'}, 'cash': {'totle_money': '0', 'will_brokerage': '0', 'going_brokerage': '0', 'finish_brokerage': '0', 'cash': '0', 'card': '0'}, 'shopping': {'totle_money': '0', 'will_brokerage': '0', 'going_brokerage': '0', 'finish_brokerage': '0', 'cash': '0', 'card': '0'}})
    # DBSCLISTDESCRESULT=["noPaidCommissionAmount","subjectPendingAmount","paidingCommissionAmount","paidCommissionAmount","noPaidCashAmount","noPaidShoppingCardAmount"]
    def myCommissionAssert(self, driver, env="t"):
        driver = self.myCommission(driver)
        logger.info(driver)
        sqlList = DBsqlCommissionMine.DBSCLIST
        dataSql = MysqlSetting(env=env).selectDataSqlMine(sqlList)
        logger.info(dataSql)
        resultdbc = DBsqlCommissionMine.DBSCLISTDESCRESULT
        resultdbcDic = {}
        listImput = []
        flagIndex = False
        resultdbcDicResultCompare = {}
        for k, v in driver[1].items():
            resultdbcIndex = 0
            for kk, vv in v.items():
                logger.info(type(vv))
                listImput.append(k)
                # if isinstance(vv,str):
                #     vv = decimal.Decimal(vv)
                if resultdbcIndex == 1:
                    try:
                        resultdbcDic[k]["historyWait"] = vv
                    except:
                        resultdbcDic[k] = {}
                        resultdbcDic[k]["historyWait"] = vv
                else:
                    try:
                        resultdbcDic[k][resultdbc[resultdbcIndex]] = vv
                    except:
                        resultdbcDic[k] = {}
                        resultdbcDic[k][resultdbc[resultdbcIndex]] = vv
                resultdbcIndex += 1
        logger.info(resultdbcDic)
        for k, v in resultdbcDic.items():
            resultdbcDicResultCompare[k] = {}
            resultdbcDicResultCompare[k]["resultDetail"] = {}
            for kk, vv in v.items():
                if kk != "historyWait":
                    dataSqlV = str(dataSql[kk])
                    # vv = str
                    vv = vv.replace(',', '')
                    logger.info(dataSqlV)
                    logger.info(vv)
                    logger.info(type(vv))
                    logger.info(type(dataSqlV))
                    try:
                        if dataSqlV == vv:
                            resultdbcDicResultCompare[k][kk] = "success"
                        else:
                            resultdbcDicResultCompare[k][kk] = "fail"
                            resultdbcDicResultCompare[k]["result"] = "fail"
                            resultdbcDicResultCompare[k]["resultDetail"][kk] = {"db": dataSqlV, "ui": vv}
                    except:
                        resultdbcDicResultCompare[k] = {}
                        if dataSqlV == vv:
                            resultdbcDicResultCompare[k][kk] = "success"
                        else:
                            resultdbcDicResultCompare[k][kk] = "fail"
                            resultdbcDicResultCompare[k]["result"] = "fail"
                            resultdbcDicResultCompare[k]["resultDetail"][kk] = {"db": dataSqlV, "ui": vv}
        logger.info(resultdbcDicResultCompare)
        return (driver, resultdbcDicResultCompare)
    
    # 牵头人佣金
    def lookCommissionData(self, driver):
        lookCommissionDataDic = {}
        # 进入牵头人佣金
        sleep(1)
        el6 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_channel")
        el6.click()
        dataList = ['shouldAmount', 'notAmount', 'auditAmount', 'alreadyAmount', 'estimateAmount', 'surplusAmount']
        # ['ShouldBrokerage', 'UnApplayBrokerage', 'ReViewBrokerage', 'UnAlreadyPayBrokerage', 'PredictBrokerage',
        #  'WaitBrokerage']
        # 应结佣
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvShouldBrokerage").text
        lookCommissionDataDic[dataList[0]] = text
        # 未提交
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvUnApplayBrokerage").text
        lookCommissionDataDic[dataList[1]] = text
        # 审核中
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvReViewBrokerage").text
        lookCommissionDataDic[dataList[2]] = text
        # 已到账
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvUnAlreadyPayBrokerage").text
        lookCommissionDataDic[dataList[3]] = text
        # 预计结佣
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvPredictBrokerage").text
        lookCommissionDataDic[dataList[4]] = text
        # 待结佣
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvWaitBrokerage").text
        lookCommissionDataDic[dataList[5]] = text
        # 结佣记录
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_brokerage_record")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        sleep(3)
        # 退出牵头人佣金，返回我的界面
        el7 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el7.click()
        # 返回到首页
        # 我的界面回到首页
        driver = self.mineGoHome(driver)
        for k, v in lookCommissionDataDic.items():
            v = v.replace(',', '')
            v = Decimal(v).quantize(Decimal('0.00'))
            lookCommissionDataDic[k] = v
        return (driver, lookCommissionDataDic)
    
    # 牵头人佣金校验
    def initiatorCommissionAssert(self, driver):
        driver = self.lookCommissionData(driver)
        ms = MysqlSetting()
        sqlList = SQLInitiator.DBSCLIST
        dbData = ms.selectDataSqlInitiator(sqlList)
        # 2020-12-14 16:59:51,366 broker_spider.py [line:515] INFO {'ShouldBrokerage': '7,936,056.74', 'UnApplayBrokerage': '7,896,056.74', 'ReViewBrokerage': '40,000.00', 'UnAlreadyPayBrokerage': '0.00', 'PredictBrokerage': '10,021,460.80', 'WaitBrokerage': '2,085,404.06'}
        # 2020-12-14 16:59:51,366 broker_spider.py [line:516] INFO {'notAmount': Decimal('7896056.74'), 'shouldAmount': Decimal('7936056.74'), 'auditAmount': Decimal('40000.00'), 'alreadyAmount': Decimal('0.00'), 'estimateAmount': Decimal('10021460.80'), 'surplusAmount': Decimal('17957517.54')}
        logger.info(driver[1])
        logger.info(dbData)
        # ["notAmount", "shouldAmount", "auditAmount",
        #                           "alreadyAmount", "estimateAmount", "surplusAmount"]
        resultdbc = SQLInitiator.DBSCLISTDESCRESULT
        resultdbcIndex = 0
        initiatorCommissionAssertDic = {}
        # Decimal('7936056.74').quantize(Decimal('0.00'))
        initiatorCommissionAssertDic["failDetail"] = {}
        for k, v in driver[1].items():
            initiatorCommissionAssertDic[k] = {}
            if dbData[k] == v:
                initiatorCommissionAssertDic[k] = {"result": "Success", "detail": {"expect": dbData[k], "actual": v}}
            else:
                initiatorCommissionAssertDic[k] = {"result": "fail", "detail": {"expect": dbData[k], "actual": v}}
                initiatorCommissionAssertDic["result"] = "Fail"
                initiatorCommissionAssertDic["failDetail"][k] = {"expect": dbData[k], "actual": v}
            # resultdbcIndex+=1
        logger.info(initiatorCommissionAssertDic)
        return (driver[0], initiatorCommissionAssertDic)
    
    # 查看结佣记录(在进入我的佣金已查看)
    def lookCommissionRecord(self):
        pass
    
    # 企业认证登记
    def companyAuthencition(self, driver, xinTax, phone):
        ta = ToolsAppium()
        if False == ta.is_element_exist(driver, ParamIdApp.USEAPP + ":id/llWite"):
            driver = self.homeGoMine(driver)
        driver = self.assertLogin(driver)
        driver = self.loginApp(driver, phone)
        driver = self.companyIdentify(driver, xinTax)
        driver = self.bankAccount(driver[0], driver[1])
        driver = self.receiveFund(driver[0], driver[1])
        driver = self.administratorEntry(driver[0], driver[1])
        return driver
    
    # 管理员信息录入
    def administratorEntry(self, driver, companyIdentifyDic):
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgIDCardFront")
        el1.click()
        sleep(3)
        if ToolsAppium().is_element_exist(driver, element='拒绝'):
            el1 = driver.find_element_by_id("com.android.packageinstaller:id/permission_allow_button")
            el1.click()
            sleep(3)
        el2 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[10]/android.view.View")
        el2.click()
        sleep(3)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el3.click()
        sleep(3)
        el4 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgIDCardSide")
        el4.click()
        sleep(3)
        el5 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[12]/android.view.View")
        el5.click()
        sleep(3)
        el6 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el6.click()
        el7 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOwnerName")
        # 输入管理员姓名
        companyIdentifyDic["admin_id_card_name"] = "薛业乔"
        el7.send_keys("薛业乔")
        el8 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOwnerIdCardNum")
        companyIdentifyDic["admin_id_card_num"] = "44010319900822244x"
        el8.send_keys("44010319900822244x")
        # 滑动界面
        ta = ToolsAppium()
        driver = ta.swipeUp(driver)
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOwnerIDCardValidDate")
        el1.click()
        # 证件有效期
        driver = ta.legalSwipeUp(driver)
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOwnerPhone")
        # 管理员手机号
        companyIdentifyDic["admin_phone"] = "15700000006"
        el1.send_keys("15700000006")
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOwnerEmail")
        # 管理员手机号
        companyIdentifyDic["admin_email"] = "15700000006@163.com"
        el2.send_keys("15700000006@163.com")
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgAttorney")
        el3.click()
        sleep(3)
        el4 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[9]/android.view.View")
        el4.click()
        sleep(3)
        el5 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el5.click()
        el6 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el6.click()
        sleep(3)
        text = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvResult").text
        if "企业认证已完成" in text:
            logger.info("企业认证完成")
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBack")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabMine")
        el2.click()
        sleep(3)
        return (driver, companyIdentifyDic)
    
    # 收款金额验证
    def receiveFund(self, driver, companyIdentifyDic):
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edMoney")
        # 输入校验金额
        companyIdentifyDic["receiveFund"] = "0.11"
        el2.send_keys("11")
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el3.click()
        sleep(3)
        return (driver, companyIdentifyDic)
    
    # 企业认证企业信息录入
    def companyIdentify(self, driver, xinTax='91110000100000489L', companyIdentifyDic={}):
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvAuthState")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btnApply")
        el2.click()
        sleep(3)
        # ToolsAppium().waitWait(driver, idContent='com.easylife.house.broker.test:id/imgBusinessv')
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgBusiness")
        el3.click()
        sleep(3)
        if ToolsAppium().is_element_exist(driver, element='拒绝'):
            el4 = driver.find_element_by_id("com.android.packageinstaller:id/permission_allow_button")
            el4.click()
            sleep(3)
        el5 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[9]/android.view.View")
        el5.click()
        sleep(3)
        el6 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el6.click()
        sleep(3)
        el7 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edIDCardNum")
        el7.clear()
        sleep(3)
        # 输入企业信用代码
        el7.send_keys(xinTax)
        companyIdentifyDic["social_credit_code"] = xinTax
        el8 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el8.click()
        sleep(3)
        el9 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgBusiness")
        el9.click()
        sleep(3)
        el10 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[6]/android.view.View")
        el10.click()
        sleep(3)
        el11 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el11.click()
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edCompanyName")
        el1.click()
        # 输入公司名称
        el1.send_keys("公司名称")
        driver.hide_keyboard()
        companyIdentifyDic['company_name'] = "公司名称"
        sleep(3)
        el12 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edIDCardNum")
        el12.clear()
        sleep(3)
        el12.send_keys(xinTax)
        # 滑动界面
        ta = ToolsAppium()
        driver = ta.swipeUp(driver)
        #
        driver = ta.dateSwipeUp(driver, companyIdentifyDic)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edAddress")
        el1.click()
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edAddress")
        # 获取信息地址
        companyIdentifyDic["company_address"] = "北京朝阳区"
        el2.send_keys("北京朝阳区")
        driver.hide_keyboard()
        sleep(3)
        ta.industrySwipeUp(driver)
        ta.scaleSwipeUp(driver)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edRegisterMoney")
        el1.click()
        el1.clear()
        # 输入注册资本
        companyIdentifyDic["registered_capital"] = "100000"
        el1.send_keys("100000")
        driver.hide_keyboard()
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edBusinessScope")
        el1.clear()
        # 输入经营范围
        companyIdentifyDic["biz_scope"] = "计算机技术"
        el1.send_keys("计算机技术")
        # 法人身份证上传，选择日期暂未处理
        driver = ta.swipeUp(driver)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgIDCardFront")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[6]/android.view.View")
        el2.click()
        sleep(3)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el3.click()
        sleep(3)
        el4 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgIDCardSide")
        el4.click()
        sleep(3)
        el5 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[9]/android.view.View")
        el5.click()
        sleep(3)
        el6 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el6.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edOwnerName")
        el2.clear()
        # 法人姓名
        el2.send_keys("薛业乔")
        companyIdentifyDic["corporation_id_card_name"] = "薛业乔"
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edOwnerIdCardNum")
        # 法人身份证号
        companyIdentifyDic["corporation_id_card_num"] = "44010319900822244x"
        el3.send_keys("44010319900822244x")
        el4 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edOwnerPhone")
        # 法人手机号
        companyIdentifyDic["corporation_phone"] = "15700000012"
        el4.send_keys("15700000012")
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOwnerIDCardValidDate")
        el1.click()
        ta.legalSwipeUp(driver)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        return (driver, companyIdentifyDic)
    
    # 银行账号录入
    def bankAccount(self, driver, companyIdentifyDic):
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/imgAddBankCard")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[6]/android.view.View")
        el2.click()
        sleep(3)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/button_apply")
        el3.click()
        sleep(3)
        el4 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edBankCardOwnerName")
        # 录入银行账户
        companyIdentifyDic["account_name"] = "工商银行"
        el4.send_keys("工商银行")
        sleep(3)
        el5 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edBankCardNum")
        # 录入银行账号
        companyIdentifyDic["account_num"] = "12323234"
        el5.send_keys("12323234")
        sleep(1)
        # 滑动界面
        ta = ToolsAppium()
        driver = ta.swipeUp(driver)
        # 暂时不取出来信息
        driver = ta.bankBranchChoose(driver)
        driver = ta.areaChoose(driver)
        driver = ta.bankChoose(driver)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        return (driver, companyIdentifyDic)
    
    # 企业认证信息查询
    def companyAuthencitionInfomation(self, driver):
        driver = self.enterpriseCertification(driver)
        driver = self.bankAccountInfomation(driver[0], driver[1])
        driver = self.administratorInfomation(driver[0], driver[1])
        return driver
    
    # 企业认证信息查询
    def enterpriseCertification(self, driver):
        companyIdentifyDic = {}
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/rl_company_auth")
        el1.click()
        el2 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.LinearLayout/android.widget.RelativeLayout/android.widget.ImageView")
        el2.click()
        sleep(3)
        # 公司名称
        companyIdentifyDic["company_name"] = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvCompanyName").text
        # 认证状态
        companyIdentifyDic["base_status"] = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvAuthState").text
        # 证件类型
        companyIdentifyDic["certificate_type"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvIDCardType").text
        # 信用代码
        companyIdentifyDic["social_credit_code"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvIDCardNum").text
        # 所在省市
        companyIdentifyDic["province_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvProvince").text
        # 详细地址
        companyIdentifyDic["company_address"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvAddress").text
        # 所在行业
        companyIdentifyDic["industry_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvIndustry").text
        # 企业规模
        companyIdentifyDic["customer_scale"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvCompanyScale").text
        # 注册资本
        companyIdentifyDic["biz_scope"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvRegisterMoney").text
        # 经营范围
        companyIdentifyDic["biz_scope"] = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBusinessScope").text
        # 滑动界面
        ta = ToolsAppium()
        driver = ta.swipeUp(driver)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el1.click()
        sleep(3)
        el1 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.RelativeLayout/android.widget.ImageView")
        el1.click()
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el1.click()
        sleep(3)
        el1 = driver.find_element_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[3]/android.widget.LinearLayout/android.widget.RelativeLayout/android.widget.ImageView")
        el1.click()
        sleep(3)
        # 法人姓名
        companyIdentifyDic["corporation_id_card_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerName").text
        # 认证状态
        companyIdentifyDic["authenticate_status"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerAuthState").text
        # 证件类型
        companyIdentifyDic["corporation_id_card_type"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerIDCardType").text
        # 证件号码
        companyIdentifyDic["corporation_id_card_num"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerIDCardNum").text
        # 法人证件有效期
        companyIdentifyDic["corporation_id_card_expiry_date"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerIDCardValidDate").text
        # 手机号
        companyIdentifyDic["corporation_phone"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerPhone").text
        # 返回到我的界面
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el1.click()
        sleep(3)
        return (driver, companyIdentifyDic)
    
    # 银行账号信息查询
    def bankAccountInfomation(self, driver, companyIdentifyDic):
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/layAccountInfo")
        el1.click()
        sleep(3)
        # 账号名称
        companyIdentifyDic["account_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvBankCardOwnerName").text
        # 认证状态
        companyIdentifyDic["bank_authenticate_status"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvAuthState").text
        # 开户账号
        companyIdentifyDic["account_num"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvBankCardNum").text
        # 开户行
        companyIdentifyDic["bank_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvBankCardOpenBankName").text
        # 开户地区
        companyIdentifyDic["city_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvBankCardOpenBankAddress").text
        # 开户行支行
        companyIdentifyDic["bank_branch_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvBankCardOpenBankSubName").text
        # 是否申请闪佣
        companyIdentifyDic["used_ebill_flag"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvFastCommission").text
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        return (driver, companyIdentifyDic)
    
    # 管理员认证信息
    def administratorInfomation(self, driver, companyIdentifyDic):
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/layManagerInfo")
        el1.click()
        sleep(3)
        # 姓名
        companyIdentifyDic["admin_id_card_name"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerName").text
        # 认证状态
        companyIdentifyDic["admin_status"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerAuthState").text
        # 证件类型
        companyIdentifyDic["admin_id_card_type"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerIDCardType").text
        # 证件号码
        companyIdentifyDic["admin_id_card_num"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerIDCardNum").text
        # 有效期
        companyIdentifyDic["admin_id_card_expiry_date"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerIDCardValidDate").text
        # 手机号
        companyIdentifyDic["admin_phone"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerPhone").text
        # 邮箱
        companyIdentifyDic["admin_email"] = driver.find_element_by_id(
            ParamIdApp.USEAPP + ":id/tvOwnerEmail").text
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el2.click()
        sleep(3)
        # 返回到我的界面
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/action_bar_left_btn")
        el1.click()
        sleep(3)
        return (driver, companyIdentifyDic)
    
    # 企业认证获取输入值获取展示值，进行校验对比
    def compareEnterpriseCertification(self, driver, xinTax, phone):
        driver = self.companyAuthencition(driver, xinTax, phone)
        logger.info(driver)
        ta = ToolsAppium()
        if ta.is_element_exist(driver[0], ParamIdApp.USEAPP + ":id/tvIndexLive"):
            driver = self.homeGoMine(driver[0])
        driverInfo = self.companyAuthencitionInfomation(driver[0])
        logger.info(driverInfo)
        ta = ToolsAppium()
        result = ta.assertExpectDic(driver[1], driverInfo[1])
        if ta.is_element_exist(driverInfo[0], ParamIdApp.USEAPP + ":id/layCommissionWait"):
            driver = self.mineGoHome(driverInfo[0])
        driver = self.assertLogin(driver)
        return (driver, result)
    
    # 认证中,银行卡
    def rezhen(self, driver):
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/rl_company_auth")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btnApply")
        el2.click()
        sleep(3)
        # 滑动界面
        ta = ToolsAppium()
        driver = ta.swipeUp(driver)
        driver = ta.bankBranchChoose(driver)
        driver = ta.areaChoose(driver)
        driver = ta.bankChoose(driver)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        # el1 = driver.find_element_by_id(ParamIdApp.USEAPP +":id/tvBankCardOpenBankName")
        # el1.click()
        # el2 = driver.find_element_by_id(ParamIdApp.USEAPP +":id/edSearch")
        # el2.click()
        # el3 = driver.find_element_by_id(ParamIdApp.USEAPP +":id/edSearch")
        # el3.click()
        # el3.send_keys("工")
        # # ToolsAppium().keyCode(type='Sougou')
        # sleep(3)
        # driver.keyevent(66)
    
    # 认证3，管理员
    def rezhen3(self, driver):
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/rl_company_auth")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btnApply")
        el2.click()
        sleep(3)
        driver = self.administratorEntry(driver)
    
    # 点击始终允许
    def always(self):
        alwaysbutton = self.driver.find_element_by_id("com.android.packageinstaller:id/permission_allow_button")
        alwaysbutton.click()
        self.driver.implicitly_wait(30)
        alwaysbutton = self.driver.find_element_by_id("com.android.packageinstaller:id/permission_allow_button")
        alwaysbutton.click()
        # self.tm.WebDriverWaitW(driverDR=alwaysbutton,needFind="yes",someId=ParamIdApp.USEAPP +":id/tvNewHouse")
        self.driver.implicitly_wait(10)
        # alwaysbutton = self.driver.find_element_by_class_name("android.widget.TextView")
        # 新房点击
        # alwaysbutton = self.driver.find_element_by_id(ParamIdApp.USEAPP +":id/tvNewHouse")
        # 客户点击
        alwaysbutton = self.driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabCustomer")
        logger.info(alwaysbutton.text)
        alwaysbutton.click()
        self.driver.implicitly_wait(30)
        alwaysbutton = self.driver.find_elements_by_xpath(
            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout")
        xml1 = self.drier.page_source
        logger.info(xml1)
        xml2 = self.driver.find_element_by_css_selector("android.widget.EditText")
        print(xml2)

    # 企业认证判断是否登录，已经登录需要退出
    def assertLogin(self, driver):
        ta = ToolsAppium()
        if ta.is_element_exist(driver, ParamIdApp.USEAPP + ":id/tvLoginTip"):
            return driver
        else:
            el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabMine")
            el1.click()
            sleep(2)
            el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tv_mine_setting")
            el2.click()
            sleep(2)
            el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btLogin")
            el3.click()
            sleep(2)
            el4 = driver.find_element_by_id("android:id/button1")
            el4.click()
            sleep(2)
            el5 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tabIndex")
            el5.click()
            sleep(2)
            return driver
# 获取Excel用例
# 报告写入Excel
# 封装工具
class ToolsAppium:
    def __init__(self):
        pass
    
    from appium import webdriver
    import time
    '''
    #封装翻页类
    class SwipepMethod:
        def __init__(self,driver=None):
            self.pm_size=self.swipSize()

        #获取屏幕尺寸
        def swipSize(self):
            return driver.get_window_size()
    '''
    
    # 判断元素是否存在
    def isElement(self, identifyBy, c, driver):
        '''
        Determine whether elements exist
        Usage:
        isElement(By.XPATH,"//a")
        '''
        time.sleep(1)
        flag = None
        try:
            if identifyBy == "id":
                # self.driver.implicitly_wait(60)
                driver.find_element_by_id(c)
            elif identifyBy == "xpath":
                # self.driver.implicitly_wait(60)
                driver.find_element_by_xpath(c)
            elif identifyBy == "class":
                driver.find_element_by_class_name(c)
            elif identifyBy == "link text":
                driver.find_element_by_link_text(c)
            elif identifyBy == "partial link text":
                driver.find_element_by_partial_link_text(c)
            elif identifyBy == "name":
                driver.find_element_by_name(c)
            elif identifyBy == "tag name":
                driver.find_element_by_tag_name(c)
            elif identifyBy == "css selector":
                driver.find_element_by_css_selector(c)
            flag = True
        except:
            flag = False
        finally:
            return flag
    
    # 1、 从数据库读取数据存成待使用数据json
    # {'FAIL': {'id': 417, 'organization_id': 438, 'used_ebill_flag': 1}, 'SUC': {}, 'expectDic': {'id': '417', 'organization_id': '438', 'used_ebill_flag': '1'}, 'actule': {'id': 417, 'organization_id': 438, 'used_ebill_flag': 1}}
    # 读取strDbKey 预期结果字段和值，strDbValue协助定位字段和值，返回对比结果
    def getDataExcept(self, env, strDbKey, strDbValue):
        ms = MysqlSetting(env=env)
        # strDbKey="id=417!_organization_id=438!_used_ebill_flag=1"
        # strDbValue="account_num=1231255!_account_name=北京!_tableName=easylife_broker_corporation_organization_bank"
        resultData = ms.selectDataBind(strDbKey, strDbValue)
        return resultData
    
    # 解析校验数据
    def assertExpect(self, expectList, driver):
        sleep(3)
        expectListDic = {}
        itemResult = True
        driverData = driver.page_source
        for item in expectList:
            if item in driverData:
                expectListDic[expectList][item] = "Success"
            else:
                expectListDic[expectList][item]['result'] = "Fail"
                # {"driver":driver,"filename":filename}
                screenShot = self.take_screenShot(driver, item)
                driver = screenShot["driver"]
                expectListDic[expectList][item]['filepath'] = screenShot["filename"]
                itemResult = False
        if itemResult == True:
            expectListDic[expectList]["result"] = "Success"
        else:
            expectListDic[expectList]["result"] = "Fail"
        return expectListDic
    
    # 对比预期结果中的值与实际值是否一致
    def assertExpectDic(self, expectInputDic, expectOutputDic):
        expectListDic = {}
        expectListDic["failDetail"] = {}
        itemResult = True
        for k, v in expectInputDic.items():
            try:
                if expectOutputDic[k] == v:
                    expectListDic[k] = "Success"
                else:
                    expectListDic[k]["result"] = "Fail"
                    expectListDic["failDetail"][k] = {"expect": v, "actual": expectOutputDic[k]}
                    itemResult = False
            except:
                logger.info("expectOutputDic " + k + "校验字段未采集")
        if itemResult == True:
            expectListDic["result"] = "Success"
        else:
            expectListDic["result"] = "Fail"
        return expectListDic
    
    # 截图工具
    def take_screenShot(self, driver, name="takeShot"):
        '''

        :param name:
        :return:
        ''''''
        method
        explain: 获取当前屏幕的截图
        parameter
        explain：【name】 截图的名称
        Usage:
        device.take_screenShot(u"个人主页")  # 实际截图保存的结果为：2018-01-13_17_10_58_个人主页.png
        '''
        day = time.strftime("%Y-%m-%d", time.localtime(time.time()))
        fq = "./screenShots/" + day
        # fq =os.getcwd()[:-4] +‘screenShots\\‘+day    根据获取的路径，然后截取路径保存到自己想存放的目录下
        tm = time.strftime("%Y-%m-%d_%H_%M_%S", time.localtime(time.time()))
        type = '.png'
        filename = ""
        if os.path.exists(fq):
            filename = fq + "/" + tm + "_" + name + type
        else:
            os.makedirs(fq)
            filename = fq + "/" + tm + "_" + name + type
        # c = os.getcwd()
        # r"\\".join(c.split("\\"))     #此2行注销实现的功能为将路径中的\替换为\\
        driver.get_screenshot_as_file(filename)
        return {"driver": driver, "filename": filename}
    
    # 判断元素是否存在
    def is_element_exist(self, driver, element):
        source = driver.page_source
        print(source)
        if element in source:
            return True
        else:
            return False
    
    # 判断是否在首页，没有在首页
    # 选择开户地区
    def areaChoose(self, driver):
        sleep(3)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBankCardOpenBankAddress")
        el1.click()
        eleArea = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheelProvince")
        # 滑动5分
        driver = self.swipeUp1(driver, ele=eleArea, t=1000, n=1, startRate=0.4, endRate=0.25)
        eleAreaCity = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheelCity")
        driver = self.swipeUp1(driver, ele=eleAreaCity, t=1000, n=1, startRate=0.4, endRate=0.25)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOk")
        el1.click()
        sleep(3)
        return driver
    
    # 开户行支行选择
    def bankChoose(self, driver):
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBankCardOpenBankSubName")
        el2.click()
        sleep(5)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edSearch")
        el3.click()
        sleep(5)
        el4 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edSearch")
        el4.send_keys("工")
        sleep(3)
        # self.keyCode(type='Sougou')
        # el1 = driver.find_element_by_id(ParamIdApp.USEAPP +":id/edSearch")
        # el1.click()
        driver.keyevent(66)
        driver.hide_keyboard()
        el5 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvName")
        el5.click()
        sleep(3)
        return driver
    
    # 开户行选择
    def bankBranchChoose(self, driver):
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvBankCardOpenBankName")
        el1.click()
        sleep(3)
        el2 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edSearch")
        el2.click()
        sleep(3)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/edSearch")
        el3.send_keys("工")
        sleep(3)
        # self.keyCode(type='Sougou')
        # el1 = driver.find_element_by_id(ParamIdApp.USEAPP +":id/edSearch")
        # el1.click()
        # driver.keyevent(66)
        driver.press_keycode(66)
        driver.hide_keyboard()
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvName")
        el1.click()
        sleep(3)
        return driver
    
    # 企业规模（7分）
    def scaleSwipeUp(self, driver):
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvCompanyScale")
        el1.click()
        sleep(3)
        eleScale = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheeView_center")
        driver = self.swipeUp1(driver, ele=eleScale, t=1000, n=1)
        el3 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvok")
        el3.click()
        sleep(3)
        return driver
    
    # 法人信息日期控件
    def legalSwipeUp(self, driver):
        # el1 = driver.find_element_by_id(ParamIdApp.USEAPP +":id/tvOwnerIDCardValidDate")
        # el1.click()
        eleyear = driver.find_elements_by_id('com.easylife.house.broker.test:id/year')[0]  # 当前年份元素
        driver = self.swipeUp1(driver, ele=eleyear, t=1000, n=1)
        elemonth = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/month")
        driver = self.swipeUp1(driver, ele=elemonth, t=1000, n=1)
        eleday = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/day")
        driver = self.swipeUp1(driver, ele=eleday, t=1000, n=1)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btnSubmit")
        el1.click()
        sleep(3)
        return driver
    
    # 所属行业滑动选择（5分）
    def industrySwipeUp(self, driver):
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvIndustry")
        el1.click()
        sleep(3)
        eleIndustry = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheeView_center")
        driver = self.swipeUp1(driver, ele=eleIndustry, t=1000, n=1, startRate=0.4, endRate=0.25)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        eleIndustry = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheeView_center")
        driver = self.swipeUp1(driver, ele=eleIndustry, t=1000, n=1, startRate=0.4, endRate=0.25)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        eleIndustry = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheeView_center")
        driver = self.swipeUp1(driver, ele=eleIndustry, t=1000, n=1, startRate=0.4, endRate=0.25)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        eleIndustry = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheeView_center")
        driver = self.swipeUp1(driver, ele=eleIndustry, t=1000, n=1, startRate=0.4, endRate=0.25)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvSubmit")
        el1.click()
        sleep(3)
        return driver
    
    # 通过抓取到时间控件的元素滑动选择日期
    def dateSwipeUp(self, driver, companyIdentifyDic):
        # 选择日期
        driver.find_element_by_id('com.easylife.house.broker.test:id/tvCompanyValidityDate').click()  # 点击日期选择
        eleyear = driver.find_elements_by_id('com.easylife.house.broker.test:id/year')[0]  # 当前年份元素
        print(eleyear)
        driver = self.swipeUp1(driver, ele=eleyear, t=1000, n=1)
        elemonth = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/month")
        driver = self.swipeUp1(driver, ele=elemonth, t=1000, n=1)
        eleday = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/day")
        driver = self.swipeUp1(driver, ele=eleday, t=1000, n=1)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/btnSubmit")
        el1.click()
        # 选择城市
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvProvince")
        el1.click()
        sleep(3)
        elePrivince = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/wheelProvince")
        driver = self.swipeUp1(driver, ele=elePrivince, t=1000, n=1, startRate=0.4, endRate=0.25)
        el1 = driver.find_element_by_id(ParamIdApp.USEAPP + ":id/tvOk")
        el1.click()
        return driver
    
    # 调用输入法解决输入
    def keyCode(self, type='Sougou'):
        command2 = 'adb shell ime set com.sohu.inputmethod.sogou/.SogouIME'
        command3 = 'adb shell ime set io.appium.android.ime/.UnicodeIME'
        
        # 列出系统现在所安装的所有输入法 os.system(command0)
        # 打印系统当前默认的输入法 os.system(command1)
        # 切换华为输入法为当前输入法 os.system(command2)
        # 切换appnium输入法为当前输入法 os.system(command3)
        if type == 'Sougou':
            os.system(command2)
        elif type == 'Appium':
            os.system(command3)
    
    # 向上滑动屏幕
    def swipeUp(self, driver, t=1000, n=1):
        l = driver.get_window_size()
        x1 = l['width'] * 0.5  # x坐标
        y1 = l['height'] * 0.75  # 起始y坐标
        y2 = l['height'] * 0.20  # 终点y坐标
        for i in range(n):
            driver.swipe(x1, y1, x1, y2, t)
        return driver
        # 向上滑动屏幕
    
    def swipeUp01(self, driver, t=1000, n=1):
        l = driver.get_window_size()
        x1 = l['width'] * 0.5  # x坐标
        y1 = l['height'] * 0.75  # 起始y坐标
        y2 = l['height'] * 0.50  # 终点y坐标
        for i in range(n):
            driver.swipe(x1, y1, x1, y2, t)
        return driver
    
    # 向上滑动屏幕(时间控件)(城市控件推荐0.40,0.25)
    def swipeUp1(self, driver, ele=None, t=1000, n=1, startRate=0.35, endRate=0.25):
        loctionD = ele.location
        sizeD = ele.size
        sizeDD = sizeD['width'] * 0.5  # x坐标
        x1 = loctionD["x"] + sizeDD
        y1 = loctionD['y'] + sizeD['height'] * startRate  # 起始y坐标
        y2 = loctionD['y'] + sizeD['height'] * endRate  # 终点y坐标
        for i in range(n):
            driver.swipe(x1, y1, x1, y2, t)
        return driver
    
    # 向下滑动屏幕
    def swipeDown(self, driver, t=500, n=1):
        l = driver.get_window_size()
        x1 = l['width'] * 0.5  # x坐标
        y1 = l['height'] * 0.25  # 起始y坐标
        y2 = l['height'] * 0.75  # 终点y坐标
        for i in range(n):
            driver.swipe(x1, y1, x1, y2, t)
        return driver
    
    # 向左滑动屏幕
    def swipeLeft(self, driver, t=500, n=1):
        l = driver.get_window_size()
        x1 = l['width'] * 0.75
        y1 = l['height'] * 0.5
        x2 = l['width'] * 0.05
        for i in range(n):
            driver.swipe(x1, y1, x2, y1, t)
        return driver
    
    # 向右滑动屏幕
    def swipRight(self, driver, t=500, n=1):
        l = driver.get_window_size()
        x1 = l['width'] * 0.05
        y1 = l['height'] * 0.5
        x2 = l['width'] * 0.75
        for i in range(n):
            driver.swipe(x1, y1, x2, y1, t)
        return driver
    
    # 显式等待
    def waitWait(self, driver, idContent):
        # 引入WebDriverWait
        from selenium.webdriver.support.ui import WebDriverWait
        # 引入expected_conditions类，并重命名为EC
        from selenium.webdriver.support import expected_conditions as EC
        # 引入By类
        from selenium.webdriver.common.by import By
        # 设置等待
        wait = WebDriverWait(driver, 10, 0.5)
        wait.until(EC.presence_of_element_located((By.ID, idContent)))
        return driver
    
    # 获取当前时间进行解析
    def getDateStr(self):
        import time
        timeListList = []
        timeStr = time.strftime("%Y-%m-%d")
        timeList = timeStr.split("-")
        for v in timeList:
            v = int(v)
            timeListList.append(v)
        return timeListList
# 断言 log
class AssertResult:
    def __init__(self):
        pass
    
    # unittest断言
    def assertUnittest(self, isUnittest="unittest", expectData=None, actualData=None):
        if isUnittest != None:
            isUnittest.assertEqual(expectData, actualData)
            return True
        else:
            return False
    
    # 自定义断言
    def assertCostom(self, isCustom="custom", expectData=None, actualData=None):
        if not isinstance(expectData, str):
            expectData = str(expectData)
        if not isinstance(actualData, str):
            actualData = str(actualData)
        if isCustom != None:
            try:
                if expectData == actualData:
                    msg = "预期 " + expectData + "实际 " + actualData + ",对比结果为相同"
                    return {"code": 1, "msg": msg}
                elif expectData != actualData:
                    msg = "预期 " + expectData + "实际 " + actualData + ",对比结果为不相同"
                    return {"code": 0, "msg": msg}
            except:
                msg = "预期 " + expectData + "实际 " + actualData + ",对比结果为不相同"
                return {"code": 0, "msg": msg}
    
    # 打印日志
    def assertLog(self, contentData):
        logger.info(contentData)
# 获取driver
class GetDriver:
    def __init__(self):
        pass
    
    # Android data
    def getAndroidData(self):
        data = {
            "platformName": "Android",
            "platformVersion": "7.0",
            "deviceName": "Galaxy S6 edge+",
            "appPackage": ParamIdApp.USEAPP + "",
            "appActivity": "com.easylife.house.broker.ui.MainActivity",
            "resetKeyboard": "true"
        }
        return data
    
    # Android setData
    def setAndroidData(self, field, value):
        initNum = 0
        if field != None:
            for i in field:
                data = self.getAndroidData()
                data[i] = value[initNum]
                initNum += 1
        else:
            data = self.getAndroidData()
        return data
    
    # ios data
    def getIosData(self):
        data = {
            "platformName": "Android",
            "platformVersion": "7.0",
            "deviceName": "Galaxy S6 edge+",
            "appPackage": ParamIdApp.USEAPP + "",
            "appActivity": "com.easylife.house.broker.ui.MainActivity",
            "resetKeyboard": "true"
        }
        return data
    
    # ios setData
    def setIosData(self, field=None, value=None):
        initNum = 0
        if field != None:
            for i in field:
                data = self.getAndroidData()
                data[i] = value[initNum]
                initNum += 1
        else:
            data = self.getAndroidData()
        return data
    
    # 返回driver
    def chooseDriver(self, platform="ANDROID", field=None, value=None):
        if platform == "ios" or platform == "IOS":
            data = self.setIosData(field=field, value=value)
        elif platform == "android" or platform == "ANDROID":
            data = self.setIosData(field=field, value=value)
        else:
            logger.info("如果是Android平台请输入“ANDROID”，如果是ios平台请输入“IOS”")
        driver = webdriver.Remote('http://localhost:4723/wd/hub', data)
        return driver
class ExcelTool:
    def __init__(self, excelFile, sheetName=None):
        self.excelFile = excelFile
        if sheetName != None:
            self.sheetName = sheetName
        else:
            self.sheetName = "Sheet1"
        self.work_book = openpyxl.load_workbook(self.excelFile)
    
    # 获取工作表
    def getSheetValue(self):
        sheet = self.work_book[self.sheetName]
        return sheet
    
    # row = xl_sheet.max_row 获取行数
    def getMaxRows(self):
        maxNum = self.getSheetValue().max_row
        return maxNum
    
    # column = xl_sheet.max_column 获取列数
    def getMaxColumn(self):
        maxColumnNum = self.getSheetValue().max_column
        return maxColumnNum
    
    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.getSheetValue().max_row
        columns = self.getSheetValue().max_column
        return rows, columns
    
    # 获取某列的所有值
    def get_col_value(self, column, rowNum=1):
        rows = self.getSheetValue().max_row
        column_data = []
        for i in range(rowNum, rows + 1):
            cell_value = self.getSheetValue().cell(row=i, column=column).value
            if cell_value != None:
                column_data.append(cell_value)
            else:
                break
        return column_data
    
    # 获取某行所有值
    def get_row_value(self, row, columnNum=1):
        columns = self.getSheetValue().max_column
        row_data = []
        for i in range(columnNum, columns + 1):
            cell_value = self.getSheetValue().cell(row=row, column=i).value
            if cell_value != None:
                row_data.append(cell_value)
            else:
                break
        return row_data
    
    # 获取
    
    # 读取指定位置的值
    def getCellValue(self, row, column):
        cellValue = self.getSheetValue().cell(row, column).value
        return cellValue
    
    # 获取所有行
    def getAllRowDataToTuple(self):
        rows = self.getSheetValue().rows
        cases = []
        for row in rows:
            row_cases = []
            for cell in row:
                row_cases.append(cell.value)
            cases.append(tuple(row_cases))
        return cases
    
    # 向指定位置写入数据
    def writeCellValue(self, row, column, rcValue):
        cellIndex = self.getSheetValue().cell(row, column)
        if not isinstance(rcValue, str):
            rcValue = str(rcValue)
        cellIndex.value = rcValue
    
    # 向一行写入一条list
    def writeCellRow(self, list, row, column=1):
        column = column
        for i in list:
            self.writeCellValue(row=row, column=column, rcValue=i)
            column += 1
    
    # 将工作簿保存
    def saveWorkbook(self, pathFile):
        self.work_book.save(pathFile)
        self.work_book.close()
class DealExcelTool:
    def __init__(self):
        pass
    
    # 读取数据的业务ID数据存放入列表
    def readOpretion(self, idOp, startPostition, endPosition):
        pass
    
    # 获取用例文件的全路径
    def getTestFileName(self):
        testFileName = self.getFilePath() + '/' + ExcelConfig.TESTCASEALLFile
        testFileName = testFileName.replace('\\', '/')
        return testFileName
    
    # 获取报告文件的全路径
    def getReportFileName(self):
        reportFileName = self.getFileReport() + '/' + ExcelConfig.REPORTPATHFILECURRENT
        reportFileName = reportFileName.replace('\\', '/')
        return reportFileName
    
    # 获取报告文件的全路径复制前
    def getReportFilePreName(self):
        reportFileName = self.getFileReport() + '/' + ExcelConfig.REPORTPATHFILE
        reportFileName = reportFileName.replace('\\', '/')
        return reportFileName
    
    # 获取临时存储文件的全路径
    def getTempFileName(self):
        tempDBFileName = self.getProjectPath() + ExcelConfig.TEMPDBFILEPATH
        tempDBFileName = tempDBFileName.replace('\\', '/')
        return tempDBFileName
    
    # 获取mac的Firefox driver的全路径
    def getMacFirefoxDriver(self, name=None):
        if name == "firefox":
            macFirefoxDriver = self.getProjectPath() + ExcelConfig.CONFIGTOOLPATH + WebSelenium.MACFIREFOXDRIVER
        else:
            macFirefoxDriver = self.getProjectPath() + ExcelConfig.CONFIGTOOLPATH + WebSelenium.MACCHROMEDRIVER
        macFirefoxDriver = macFirefoxDriver.replace('\\', '/')
        return macFirefoxDriver
    
    # 获取windows的Firefox driver的全路径
    def getWindowsFirefoxDriver(self, name=None):
        if name == "firefox":
            windowsFirefoxDriver = self.getProjectPath() + ExcelConfig.CONFIGTOOLPATH + WebSelenium.WINDOWSFIREFOXDRIVER
        else:
            windowsFirefoxDriver = self.getProjectPath() + ExcelConfig.CONFIGTOOLPATH + WebSelenium.WINDOWSCHROMEDRIVER
        windowsFirefoxDriver = windowsFirefoxDriver.replace('\\', '/')
        return windowsFirefoxDriver
    
    # 获取用例文件的路径
    def getFilePath(self):
        testCaseFilePath = self.getProjectPath() + ExcelConfig.TESTCASEALL
        return testCaseFilePath
    
    # 获取报告文件的路径
    def getFileReport(self):
        reportFilePath = self.getProjectPath() + ExcelConfig.REPORTPATH
        return reportFilePath
    
    # 获取项目路径
    def getProjectPath(self):
        projectPath = ExcelConfig.PROJECTPATH
        root_path = os.path.abspath(os.path.dirname(__file__))
        root_path = self.changeProjectPath(root_path)
        reform = re.compile("/(.*?)/test/test_api", re.S)
        projectPath = reform.findall(root_path)[0]
        root_path = root_path.split(projectPath)[0]
        # proPath = root_path[0]
        # logger.info(proPath)
        proPath = root_path + projectPath
        return proPath
    
    # 适配不同的系统，获取项目路径
    def changeProjectPath(self, pathGet=None):
        sysPlat = platform.system()
        print(type(sysPlat))
        print(sysPlat)
        if "Windows" in sysPlat:
            print("here")
            pathGet = pathGet.replace('\\', '/')
            print(pathGet)
        return pathGet
class WebSelenium:
    WINDOWSFIREFOXDRIVER = "geckodriver.exe"
    MACFIREFOXDRIVER = "geckodriver"
    WINDOWSCHROMEDRIVER = "chromedriver84.exe"
    MACCHROMEDRIVER = "chromedriver84"
class ExcelConfig:
    PROJECTPATH = 'hopsapi_test1'
    TESTCASEALL = '/testcase/testCaseAll'
    REPORTPATH = '/report/excelReport'
    TESTCASEALLFile = 'test.xlsx'
    TESTCASEALLSHEET = 'testSheet'
    # 参数化的sheet
    PARAMETERCASESHEET = 'parameterSheet'
    REPORTPATHFILE = 'testReport.xlsx'
    REPORTPATHSHEET = 'testReport'
    # 临时存储文本
    TEMPDBFILEPATH = '/db_file/TempDB.txt'
    REPORTPATHFILECURRENT = 'testReport' + str(time.strftime('%Y-%m-%d%H%M%S', time.localtime(time.time()))) + '.xlsx'
    CONFIGTOOLPATH = "/utils/config_tool/"
# 配置测试环境
class ParamIdApp:
    TESTAPP = "com.easylife.house.broker.test"
    PRODUCTAPP = "com.easylife.house.broker"
    USEAPP = TESTAPP
class MysqlSetting:
    def __init__(self, env="t"):
        self.connect = self.getConnection(env)
        self.cursor = self.getCursor()
    
    def getConnection(self, env='t'):
        if env == "h":
            listi = DBSetting.MYSQLSETTINGH
        elif env == "t":
            listi = DBSetting.MYSQLSETTINGT
        else:
            listi = env
        host = listi["host"]
        user = listi["user"]
        password = listi["password"]
        db = listi["db"]
        port = listi["port"]
        connection = pymysql.connect(host=host,
                                     user=user,
                                     password=password,
                                     port=port,
                                     db=db,
                                     charset='utf8mb4',
                                     cursorclass=pymysql.cursors.DictCursor)
        return connection
    
    def getCursor(self):
        cursor = self.connect.cursor()
        return cursor
    
    # key 是一个元组，包括两个元组，第一个元组是键，第二个是值
    # 插入数据
    def insertData(self, *key, **kwargs):
        cursor = self.cursor
        # Create a new record
        print(key)
        key0 = str(key[0])
        num = len(key[1])
        tableName = kwargs["tableName"]
        valueData = Common().getCValue(num)
        # sql = "INSERT INTO `users` (`email`, `password`) VALUES (%s, %s)"
        sql = "INSERT INTO " + tableName + key0 + " VALUES " + valueData
        print(sql)
        print(key[1])
        # cursor.execute(sql, ('webmaster@python.org', 'very-secret'))
        cursor.execute(sql, key[1])
    
    # key 是一个元组，包括两个元组，第一个元组是键，第二个是值
    # 插入数据
    def insertData01(self, *key, **kwargs):
        cursor = self.cursor
        # Create a new record
        print(key)
        key0 = str(key[0][0])
        num = len(key[0][1])
        tableName = kwargs["tableName"]
        valueData = Common().getCValue(num)
        # sql = "INSERT INTO `users` (`email`, `password`) VALUES (%s, %s)"
        sql = "INSERT INTO " + tableName + key0 + " VALUES " + valueData
        print(sql)
        print(key[0][1])
        # cursor.execute(sql, ('webmaster@python.org', 'very-secret'))
        cursor.execute(sql, key[0][1])
    
    # 测试
    def selectD1(self):
        sqlD = "SELECT id, organization_id, used_ebill_flag FROM easylife_broker_corporation_organization_bank WHERE account_num = %s  and account_name = %s"
        # sqlD = "SELECT id, organization_id, used_ebill_flag FROM easylife_broker_corporation_organization_bank WHERE account_num = '1231255'  and account_name = '北京'"
        paramD = ('1231255', '北京')
        cursor = self.cursor
        # cursor.fetchall()
        data = cursor.execute(sqlD, paramD)
        result = cursor.fetchall()
        return result
    
    def selectD(self):
        # sqlD="SELECT id, organization_id, used_ebill_flag FROM easylife_broker_corporation_organization_bank  WHERE account_num = %s  and account_name = %s"
        # sqlD = "SELECT id, organization_id, used_ebill_flag FROM easylife_broker_corporation_organization_bank WHERE id=417"
        sqlD = "SELECT id, organization_id, used_ebill_flag FROM easylife_broker_corporation_organization_bank "
        # sqlD = "select * from broker_organization_area where id = %d" %(1)
        
        paramD = ('1231255', '北京')
        cursor = self.cursor
        # cursor.fetchall()
        # cursor.execute(sqlD)
        a = cursor.execute(
            "SELECT id, organization_id, used_ebill_flag FROM easylife_broker_corporation_organization_bank WHERE id=417")
        # data=cursor.execute(sqlD)
        # print(cursor.fetchall())
        print(a)
        result = cursor.fetchall()
        self.closeConnect()
        return result
    
    # 查询数据的内容
    # {'CASH': {'CASH': None}, 'SHOPPING': {'SHOPPING': None}, 'NEWDOCOMMISSION': {'NEWDOCOMMISSION': Decimal('2496717.27')}, 'NEWDOINGCOMMISSION': {'NEWDOINGCOMMISSION': None}, 'NEWDONECOMMISSION': {'NEWDONECOMMISSION': None}, 'OLDDOCOMMISSION': {'OLDDOCOMMISSION': Decimal('1277850.80')}, 'OLDDOINGCOMMISSION': {'OLDDOINGCOMMISSION': Decimal('300200.00')}, 'OLDDONECOMMISSION': {'OLDDONECOMMISSION': Decimal('152818.80')}}
    def selectDataSql(self, sqlList):
        resultDic = {}
        resultDicDic = {}
        resultDicDicDic = {}
        desc = DBsqlCommissionTest.DBSCLISTDESC
        descNum = 0
        for sql in sqlList:
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            resultDic[desc[descNum]] = result[0]
            descNum += 1
            logger.info(result)
            logger.info(sql)
        listData = [resultDic["NEWDOCOMMISSION"]["NEWDOCOMMISSION"], resultDic["OLDDOCOMMISSION"]["OLDDOCOMMISSION"],
                    resultDic["NEWDOINGCOMMISSION"]["NEWDOINGCOMMISSION"], resultDic["OLDDOINGCOMMISSION"][
                        "OLDDOINGCOMMISSION"], resultDic["NEWDONECOMMISSION"]["NEWDONECOMMISSION"],
                    resultDic["OLDDONECOMMISSION"][
                        "OLDDONECOMMISSION"]]
        for i in listData:
            if i == None:
                i = Decimal('0.00')
        for k, v1 in resultDic.items():
            for k, v in v1.items():
                if v == None:
                    v = Decimal('0.00')
                    resultDicDicDic[k] = v
                else:
                    resultDicDicDic[k] = v
        resultDic = resultDicDicDic
        logger.info(resultDic)
        # resultDicDic["CASH"]=resultDic["CASH"]["CASH"]
        # resultDicDic["SHOPPING"]=resultDic["SHOPPING"]["SHOPPING"]
        # resultDicDic["DOCOMMISSION"]=resultDic["NEWDOCOMMISSION"]["NEWDOCOMMISSION"]+resultDic["OLDDOCOMMISSION"]["OLDDOCOMMISSION"]
        # resultDicDic["DOINGCOMMISSION"] = resultDic["NEWDOINGCOMMISSION"]["NEWDOINGCOMMISSION"] + resultDic["OLDDOINGCOMMISSION"][
        #     "OLDDOINGCOMMISSION"]
        # resultDicDic["DONECOMMISSION"] = resultDic["NEWDONECOMMISSION"]["NEWDONECOMMISSION"] + \
        #                                   resultDic["OLDDONECOMMISSION"][
        #                                       "OLDDONECOMMISSION"]
        resultDicDic["CASH"] = resultDic["CASH"]
        resultDicDic["SHOPPING"] = resultDic["SHOPPING"]
        resultDicDic["DOCOMMISSION"] = resultDic["NEWDOCOMMISSION"] + resultDic["OLDDOCOMMISSION"]
        resultDicDic["DOINGCOMMISSION"] = resultDic["NEWDOINGCOMMISSION"] + \
                                          resultDic["OLDDOINGCOMMISSION"]
        resultDicDic["DONECOMMISSION"] = resultDic["NEWDONECOMMISSION"] + \
                                         resultDic["OLDDONECOMMISSION"]
        return resultDicDic
    
    # 我的佣金sql取值
    def selectDataSqlMine(self, sqlList):
        resultDic = {}
        resultDicDic = {}
        resultDicDicDic = {}
        desc = DBsqlCommissionMine.DBSCLISTDESC
        descNum = 0
        for sql in sqlList:
            logger.info(sql)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if len(result) == 1:
                resultDIc = result[0]
                logger.info(resultDIc)
                for k, v in resultDIc.items():
                    resultDic[desc[descNum]] = v
            else:
                resultDic[desc[descNum]] = result
            descNum += 1
            logger.info(result)
        logger.info(resultDic)
        # 2020-11-19 15:51:40,438 broker_spider.py [line:1691] INFO {'noPaidCommissionAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00)  )  ,0)': Decimal('922130.80')}, 'noPaidCommissionAmountNew': {'sum(pendingCommissionAmount)': Decimal('3033199.47')}, 'subjectPendingAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00) ) ,0)': Decimal('922130.80')}, 'subjectPendingAmountNew': {'subject_id': 5, 'subject_name': '销售佣金', 'pendingCommissionAmount': Decimal('1994968.64')}, 'paidingCommissionAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00) ),0)': Decimal('317350.00')}, 'paidingCommissionAmountNew': {'sum(paidingCommissionAmount)': Decimal('196610.00')}, 'paidCommissionAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00)  ) ,0)': Decimal('166718.80')}, 'getKingdeePaidCashAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('2600.00')}, 'getConfirmPaidShoppingCardAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('100.00')}, 'newPaidCommissionAmount': {'sum(paidCommissionAmount)': Decimal('866929.12')}, 'noPaidCashAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('63780.00')}, 'noPaidShoppingCardAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('0.00')}}
        DBD = DBsqlCommissionMine.DBSCLISTDESCRESULT
        # 待结佣
        resultDicDicDic[DBD[0]] = resultDic[desc[0]] + resultDic[desc[1]]
        # 2224新佣金科目
        for item in resultDic[desc[3]]:
            if item["subject_id"] == "5":
                resultDicDicDic[item["subject_name"]] = resultDic[desc[2]] + item["pendingCommissionAmount"]
            else:
                resultDicDicDic[item["subject_name"]] = item["pendingCommissionAmount"]
        # 申请中
        resultDicDicDic[DBD[2]] = resultDic[desc[4]] + resultDic[desc[5]]
        # 已结佣金额
        resultDicDicDic[DBD[3]] = resultDic[desc[6]] + resultDic[desc[7]] + resultDic[desc[8]] + resultDic[desc[9]]
        # 现金奖励
        resultDicDicDic[DBD[4]] = resultDic[desc[10]]
        # 购物卡奖励
        resultDicDicDic[DBD[5]] = resultDic[desc[11]]
        logger.info(resultDicDicDic)
        return resultDicDicDic
    
    # 牵头人佣金sql取值
    # 2020-11-26 14:51:48,855 broker_spider.py [line:1849] INFO {'NotAmountgetNotAmount1': Decimal('38001.20'), 'NotAmountgetStatisticsAmountnode_amount': Decimal('7898055.54'), 'NotAmountgetStatisticsAmountcash_node_amount': None, 'NotAmountgetStatisticsAmountcard_node_amount': None, 'shouldAmountgetShouldAmount': Decimal('38001.20'), 'shouldAmountnode_amount': Decimal('7898055.54'), 'auditAmountgetInAuditAmount': Decimal('0.00'), 'auditAmountnode_amount': None, 'auditAmountcash_node_amount': None, 'auditAmountcard_node_amount': None, 'alreadyAmountgetArrivalAmount': Decimal('0.00'), 'alreadyAmountnode_amount': None, 'alreadyAmountcash_node_amount': None, 'alreadyAmountcard_node_amount': None, 'estimateAmountgetEstimateAmount': Decimal('198003.20'), 'estimateAmountgetEstimateAount': Decimal('9823457.60'), 'surplusAmountgetEstimateAmount': Decimal('198003.20'), 'surplusAmountgetShouldAmount': Decimal('38001.20'), 'surplusAmountgetEstimateAount': Decimal('9823457.60'), 'surplusAmountnode_amount': Decimal('7898055.54')}
    # 2020-11-26 15:00:27,632 broker_spider.py [line:1868] INFO {'notAmount': Decimal('7936056.74'), 'shouldAmount': Decimal('7936056.74'), 'auditAmount': Decimal('0.00'), 'alreadyAmount': Decimal('0.00'), 'estimateAmount': Decimal('10021460.80'), 'surplusAmount': Decimal('17957517.54')}
    def selectDataSqlInitiator(self, sqlList):
        resultDic = {}
        resultDicDic = {}
        resultDicDicDic = {}
        desc = SQLInitiator.DBSCLISTDESC
        descNum = 0
        for sql in sqlList:
            logger.info(sql)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if len(result) == 1:
                resultDIc = result[0]
                logger.info(resultDIc)
                for k, v in resultDIc.items():
                    if v == None:
                        v = Decimal('0.00')
                    logger.info(descNum)
                    # logger.info(desc[descNum-1])
                    resultDic[desc[descNum]] = v
            else:
                resultDic[desc[descNum]] = result
            descNum += 1
            logger.info(result)
        logger.info(resultDic)
        # 2020-11-19 15:51:40,438 broker_spider.py [line:1691] INFO {'noPaidCommissionAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00)  )  ,0)': Decimal('922130.80')}, 'noPaidCommissionAmountNew': {'sum(pendingCommissionAmount)': Decimal('3033199.47')}, 'subjectPendingAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00) ) ,0)': Decimal('922130.80')}, 'subjectPendingAmountNew': {'subject_id': 5, 'subject_name': '销售佣金', 'pendingCommissionAmount': Decimal('1994968.64')}, 'paidingCommissionAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00) ),0)': Decimal('317350.00')}, 'paidingCommissionAmountNew': {'sum(paidingCommissionAmount)': Decimal('196610.00')}, 'paidCommissionAmountOld': {'IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00)  ) ,0)': Decimal('166718.80')}, 'getKingdeePaidCashAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('2600.00')}, 'getConfirmPaidShoppingCardAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('100.00')}, 'newPaidCommissionAmount': {'sum(paidCommissionAmount)': Decimal('866929.12')}, 'noPaidCashAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('63780.00')}, 'noPaidShoppingCardAmount': {'IFNULL(sum(bai.node_amount),0)': Decimal('0.00')}}
        DBD = SQLInitiator.DBSCLISTDESCRESULT
        # 未提交佣金
        resultDicDicDic[DBD[0]] = resultDic[desc[0]] + resultDic[desc[1]] + resultDic[desc[2]] + resultDic[desc[3]]
        # 应结用
        resultDicDicDic[DBD[1]] = resultDic[desc[4]] + resultDic[desc[5]]
        # 审核中auditAmount
        resultDicDicDic[DBD[2]] = resultDic[desc[7]] + resultDic[desc[8]] + resultDic[desc[9]] + resultDic[desc[6]]
        # 已到账
        resultDicDicDic[DBD[3]] = resultDic[desc[11]] + resultDic[desc[12]] + resultDic[desc[13]] + resultDic[desc[10]]
        # 预计结佣estimateAmount
        resultDicDicDic[DBD[4]] = resultDic[desc[15]] + resultDic[desc[14]]
        # 待结佣
        resultDicDicDic[DBD[5]] = resultDic[desc[16]] + resultDic[desc[17]] + resultDic[desc[18]] + resultDic[desc[19]]
        logger.info(resultDicDicDic)
        return resultDicDicDic
    
    # 查询数据
    # 传入元组，第一个字符串"(key,key)",第二个传入元组（key),第三个传入值（value），第四个传入表名，tableName=Name,
    # <class 'tuple'>: ('(id, organization_id, used_ebill_flag)', ('account_num', 'account_name'), ('1231255', '北京'))
    # get --[{'id': 2, 'code': '100', 'create_time': datetime.datetime(2020, 8, 10, 14, 8, 21)}]
    def selectData(self, *key, **kwargs):
        # Read a single record
        keyDPatterm = re.compile("\((.*?)\)")
        logger.info(key)
        keyD = keyDPatterm.findall(key[0])[0]
        keyWhere = Common().getSValue(key[1])
        # sql = "SELECT `id`, `password` FROM `users` WHERE `email`=%s"
        sql = "SELECT " + keyD + " FROM " + kwargs["tableName"] + " WHERE" + keyWhere
        logger.info(sql)
        logger.info(key[2])
        self.cursor.execute(sql, key[2])
        result = self.cursor.fetchall()
        return result
    
    # 1、 通过where定位查询到实际结果
    def selectDataBind(self, strDbKey, strDbValue):
        DicParameter = Common().expectDB(strDbKey, strDbValue)
        result = self.selectData(DicParameter["expectKey"], DicParameter["locationListKey"],
                                 DicParameter["locationListValue"], tableName=DicParameter["tableName"])
        try:
            resultD = Common().compareData(expectDic=DicParameter["expectValue"], actule=result[0])
        except:
            resultD = "数据库获取内容为空！"
        logger.info(resultD)
        return resultD
    
    # 2、 通过sql语句查询
    def commitData(self):
        self.connect.commit()
    
    def closeConnect(self):
        self.connect.close()
class Common:
    def __init__(self):
        pass
    
    # 预期结果：字段名 = 值_字段名 = 值_字段名 = 值：字段名 = 值_字段名 = 值 'id=417!_organization_id=438!_used_ebill_flag=1':'account_num=1231255!_account_name=北京'
    #  如果是int 或者浮点类型  值的前面加一个！
    def expectDB(self, strDbKey, strDbValue):
        expectDic = {}
        expectList = []
        locationListKey = []
        locationListValue = []
        # strDbList = strDb.split(":")
        strDbListExpect = strDbKey.split("_")
        if len(strDbListExpect) > 2:
            strDbListExpect = strDbKey.split("!_")
        for i in strDbListExpect:
            i = i.split("=")
            expectDic[i[0]] = i[1]
            expectList.append(i[0])
        strDbListLocation = strDbValue.split("_")
        if len(strDbListLocation) > 2:
            strDbListLocation = strDbValue.split("!_")
        for i in strDbListLocation:
            i = i.split("=")
            if i[0] != "tableName":
                locationListKey.append(i[0])
                valueD = self.transTupleToNoStr(i[1])
                locationListValue.append(valueD)
            else:
                tableName = i[1]
        locationListKey = tuple(locationListKey)
        locationListValue = tuple(locationListValue)
        expectList = tuple(expectList)
        expectList = self.transTupleToStr(expectList)
        return {"expectValue": expectDic, "expectKey": expectList, "locationListKey": locationListKey,
                "locationListValue": locationListValue, "tableName": tableName}
    
    def transTupleToNoStr(self, keyD):
        if "!" in keyD:
            keyD = keyD[1:]
            if "." in keyD:
                keyD = float(keyD)
            else:
                keyD = int(keyD)
        return keyD
    
    def transTupleToStr(self, liti):
        liti = str(liti)
        liti = liti.replace("'", "")
        liti = liti.replace("'", "")
        liti = liti.replace('"', "")
        return liti
    
    # 字典数据对比
    def compareData(self, expectDic, actule):
        compareResult = {}
        compareResult["FAIL"] = {}
        compareResult["SUC"] = {}
        compareResult["expectDic"] = expectDic
        compareResult["actule"] = actule
        try:
            for k, v in expectDic.items():
                if actule[k] == v:
                    compareResult["SUC"][k] = "PASS"
                else:
                    compareResult["FAIL"][k] = actule[k]
        except:
            compareResult["FAIL"] = "FAIL"
        return compareResult
    
    # 生成（%s,%s）
    def getCValue(self, num):
        numList = []
        for i in range(num):
            numList.append(1)
        numListStr = str(tuple(numList))
        numListStr = numListStr.replace("1", "%s")
        return numListStr
    
    # 生成键=值  键=%s
    def getSValue(self, listTu):
        keyWhere = ""
        for key in listTu:
            keyWhere = keyWhere + " and " + key + " = %s "
        # anda = % sandb = % s
        keyWhere = keyWhere[4:]
        return keyWhere
# config配置
class DBSetting:
    MYSQLSETTINGH = {"host": "rm-2zeh739lme9f9hr08eo.mysql.rds.aliyuncs.com", "user": "easylife",
                     "password": "root123HOPSON", "db": "easylife", "port": 3306, "ssl": {'ssl': {}}}
    MYSQLSETTINGT = {"host": "124.127.103.190", "user": "root", "password": "root123HOPSON", "db": "easylife",
                     "port": 40003}
class DBsqlCommissionTest:
    CASH = "select sum(ba.node_amount) CASH from brokerage_award_info ba left join broker b on ba.broker_id=b.id where ba.award_type = 2 and ba.frozen_status=1 and ba.is_return=0 and b.broker_phone='19905021001'"
    SHOPPING = "select sum(ba.node_amount) SHOPPING from brokerage_award_info ba left join broker b on ba.broker_id=b.id where ba.award_type = 1 and ba.frozen_status=1 and ba.is_return=0 and b.broker_phone='19905021001'"
    NEWDOCOMMISSION = "select sum(cash_node_amount)+sum(card_node_amount) NEWDOCOMMISSION from brokerage_expendsubject_allocation_info beai left join broker b on beai.applyer_id = b.id where beai.cash_examine_state = 1 and b.broker_phone='19905021001' and beai.card_examine_state =1"
    NEWDOINGCOMMISSION = "select sum(cash_node_amount)+sum(card_node_amount) NEWDOINGCOMMISSION from brokerage_expendsubject_allocation_info beai left join broker b on beai.applyer_id = b.id where beai.cash_examine_state in (2,4) and b.broker_phone='19905021001' and beai.card_examine_state in (2,4)"
    NEWDONECOMMISSION = "select sum(cash_node_amount)+sum(card_node_amount) NEWDONECOMMISSION from brokerage_expendsubject_allocation_info beai left join broker b on beai.applyer_id = b.id where beai.cash_examine_state = 3 and b.broker_phone='19905021001' and beai.card_examine_state =3"
    OLDDOCOMMISSION = "select sum(front_node_broker_amount)+sum(after_node_broker_amount) OLDDOCOMMISSION from brokerage_info bi left join broker b on bi.broker_id=b.id where bi.broker_examine_state =1 and b.broker_phone='19905021001' and is_return=0 "
    OLDDOINGCOMMISSION = "select sum(front_node_broker_amount)+sum(after_node_broker_amount) OLDDOINGCOMMISSION from brokerage_info bi left join broker b on bi.broker_id=b.id where bi.broker_examine_state in (2,4) and b.broker_phone='19905021001' and is_return=0"
    OLDDONECOMMISSION = "select sum(front_node_broker_amount)+sum(after_node_broker_amount) OLDDONECOMMISSION from brokerage_info bi left join broker b on bi.broker_id=b.id where bi.broker_examine_state in (3) and b.broker_phone='19905021001' and is_return=0 "
    DBSCLIST = [CASH, SHOPPING, NEWDOCOMMISSION, NEWDOINGCOMMISSION, NEWDONECOMMISSION, OLDDOCOMMISSION,
                OLDDOINGCOMMISSION, OLDDONECOMMISSION]
    DBSCLISTDESC = ['CASH', 'SHOPPING', 'NEWDOCOMMISSION', 'NEWDOINGCOMMISSION', 'NEWDONECOMMISSION', 'OLDDOCOMMISSION',
                    'OLDDOINGCOMMISSION', 'OLDDONECOMMISSION']
# 我的佣金需要的sql
class DBsqlCommissionMine:
    # 待结佣 相加
    noPaidCommissionAmountOld = "SELECT IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00)  )  ,0) FROM  brokerage_info bi  LEFT JOIN easylife_order eo ON eo.id = bi.order_id WHERE bi.broker_id = 387889 and bi.brokerage_type_id = 1024 and" \
                                " (eo.brokerage_confirm_status = 2 or bi.is_mattress = 2) and bi.broker_examine_state = 1 and ( CASE WHEN bi.node_type = 1 THEN 1 = 1 WHEN bi.node_type = 2 THEN bi.is_completed = 1 END )"
    noPaidCommissionAmountNew = "select sum(pendingCommissionAmount) from( select info.subject_id,info.subject_name," \
                                " sum((case when info.cash_examine_state=1 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=1 then info.card_node_amount else 0 end)) pendingCommissionAmount," \
                                "sum((case when info.cash_examine_state=2 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=2 then info.card_node_amount else 0 end)) paidingCommissionAmount," \
                                " sum((case when info.cash_examine_state=3 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=3 then info.card_node_amount else 0 end)) paidCommissionAmount," \
                                "sum(case when info.cash_examine_state=1 then info.cash_node_amount else 0 end) as pendingCashAmount," \
                                "sum(case when info.cash_examine_state=2 then info.cash_node_amount else 0 end) as paidingCashAmount," \
                                "sum(case when info.cash_examine_state=3 then info.cash_node_amount else 0 end) as paidCashmount," \
                                " sum(case when info.cash_examine_state=1 then info.card_node_amount else 0 end) as pendingCardAmount," \
                                " sum(case when info.cash_examine_state=2 then info.card_node_amount else 0 end) as paidingCardAmount," \
                                " sum(case when info.cash_examine_state=3 then info.card_node_amount else 0 end) as paidCardmount" \
                                " from brokerage_expendsubject_allocation_info info" \
                                " where info.applyer_id = 387889 and info.allot_type =5 group by subject_id ) sub "
    # 224模式佣金科目  相加陈列
    subjectPendingAmountOld = "SELECT IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00) ) ,0)" \
                              " FROM brokerage_info bi LEFT JOIN easylife_order eo ON eo.id = bi.order_id WHERE bi.broker_id = 387889 and bi.brokerage_type_id = 1024 and (eo.brokerage_confirm_status = 2 or bi.is_mattress = 2) " \
                              "and bi.broker_examine_state = 1 and ( CASE WHEN bi.node_type = 1 THEN 1 = 1 WHEN bi.node_type = 2 THEN bi.is_completed = 1 END )"
    subjectPendingAmountNew = "select subject_id,subject_name, pendingCommissionAmount from(" \
                              "select info.subject_id,info.subject_name,sum((case when info.cash_examine_state=1 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=1 then info.card_node_amount else 0 end)) pendingCommissionAmount," \
                              " sum((case when info.cash_examine_state=2 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=2 then info.card_node_amount else 0 end)) paidingCommissionAmount," \
                              " sum((case when info.cash_examine_state=3 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=3 then info.card_node_amount else 0 end)) paidCommissionAmount," \
                              " sum(case when info.cash_examine_state=1 then info.cash_node_amount else 0 end) as pendingCashAmount," \
                              " sum(case when info.cash_examine_state=2 then info.cash_node_amount else 0 end) as paidingCashAmount," \
                              " sum(case when info.cash_examine_state=3 then info.cash_node_amount else 0 end) as paidCashmount," \
                              " sum(case when info.cash_examine_state=1 then info.card_node_amount else 0 end) as pendingCardAmount," \
                              "  sum(case when info.cash_examine_state=2 then info.card_node_amount else 0 end) as paidingCardAmount," \
                              " sum(case when info.cash_examine_state=3 then info.card_node_amount else 0 end) as paidCardmount" \
                              " from brokerage_expendsubject_allocation_info info " \
                              "where info.applyer_id = 387889 and info.allot_type =5" \
                              " group by subject_id ) sub"
    # 申请中 相加
    paidingCommissionAmountOld = "SELECT IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00) )" \
                                 ",0) FROM brokerage_info bi LEFT JOIN easylife_order eo ON eo.id = bi.order_id" \
                                 "  WHERE bi.broker_id = 387889 and bi.brokerage_type_id = 1024 and bi.broker_examine_state = 2" \
                                 " and ( CASE WHEN bi.node_type = 1 THEN 1 = 1 WHEN bi.node_type = 2 THEN bi.is_completed = 1 END )"
    paidingCommissionAmountNew = "select sum(paidingCommissionAmount) from( " \
                                 "select info.subject_id,info.subject_name," \
                                 " sum((case when info.cash_examine_state=1 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=1 then info.card_node_amount else 0 end)) pendingCommissionAmount," \
                                 " sum((case when info.cash_examine_state=2 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=2 then info.card_node_amount else 0 end)) paidingCommissionAmount," \
                                 " sum((case when info.cash_examine_state=3 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=3 then info.card_node_amount else 0 end)) paidCommissionAmount," \
                                 " sum(case when info.cash_examine_state=1 then info.cash_node_amount else 0 end) as pendingCashAmount," \
                                 " sum(case when info.cash_examine_state=2 then info.cash_node_amount else 0 end) as paidingCashAmount," \
                                 " sum(case when info.cash_examine_state=3 then info.cash_node_amount else 0 end) as paidCashmount," \
                                 " sum(case when info.cash_examine_state=1 then info.card_node_amount else 0 end) as pendingCardAmount," \
                                 " sum(case when info.cash_examine_state=2 then info.card_node_amount else 0 end) as paidingCardAmount," \
                                 " sum(case when info.cash_examine_state=3 then info.card_node_amount else 0 end) as paidCardmount" \
                                 " from brokerage_expendsubject_allocation_info info" \
                                 " where info.applyer_id = 387889 and info.allot_type =5" \
                                 " group by subject_id ) sub"
    
    # 已结佣金额 相加
    paidCommissionAmountOld = "SELECT IFNULL( sum( IFNULL(bi.front_node_broker_amount, 0.00) + IFNULL(bi.after_node_broker_amount, 0.00)" \
                              "  ) ,0) FROM brokerage_info bi" \
                              " LEFT JOIN brokerage_broker_apply_detail bbad ON bbad.brokerage_info_id = bi.id AND bbad.type = 1 AND bbad.model_type = 0" \
                              " LEFT JOIN brokerage_broker_apply bba ON bba.id = bbad.apply_id" \
                              " LEFT JOIN easylife_kingdee_pay_list ekpl ON ekpl.request_code = bba.code" \
                              " WHERE bi.broker_id = 387889 and bi.brokerage_type_id = 1024 and bba.examine_state = 3 and ekpl.pay_status = 3 and ( CASE WHEN node_type = 1 THEN 1 = 1 WHEN node_type = 2 THEN is_completed = 1 END )"
    getKingdeePaidCashAmount = "select IFNULL(sum(bai.node_amount),0)" \
                               " from brokerage_award_info bai" \
                               " LEFT JOIN brokerage_broker_apply_detail bbad ON bbad.brokerage_info_id = bai.id AND bbad.type = 2" \
                               " LEFT JOIN brokerage_broker_apply bba ON bba.id = bbad.apply_id LEFT JOIN easylife_kingdee_pay_list ekpl ON ekpl.request_code = bba.code WHERE" \
                               " bba.examine_state = 3 AND ekpl.pay_status = 3 AND bai.brokerage_type_id = 1024 AND bai.broker_id = 387889"
    getConfirmPaidShoppingCardAmount = "select IFNULL(sum(bai.node_amount),0)" \
                                       " from brokerage_award_info bai" \
                                       " LEFT JOIN brokerage_broker_apply_detail bbad ON bbad.brokerage_info_id = bai.id AND bbad.type = 3" \
                                       " LEFT JOIN brokerage_broker_apply bba ON bba.id = bbad.apply_id" \
                                       " LEFT JOIN shop_card_reward_apply scrcr ON scrcr.batch_code = bba.code" \
                                       " WHERE bba.examine_state in (2,3) AND scrcr.grant_status in (2,6) AND bai.broker_id = 387889" \
                                       " AND bai.brokerage_type_id = 1024"
    newPaidCommissionAmount = "select sum(paidCommissionAmount) from( select info.subject_id,info.subject_name," \
                              " sum((case when info.cash_examine_state=1 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=1 then info.card_node_amount else 0 end)) pendingCommissionAmount," \
                              " sum((case when info.cash_examine_state=2 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=2 then info.card_node_amount else 0 end)) paidingCommissionAmount," \
                              " sum((case when info.cash_examine_state=3 then info.cash_node_amount else 0 end)+(case when info.card_examine_state=3 then info.card_node_amount else 0 end)) paidCommissionAmount," \
                              " sum(case when info.cash_examine_state=1 then info.cash_node_amount else 0 end) as pendingCashAmount," \
                              " sum(case when info.cash_examine_state=2 then info.cash_node_amount else 0 end) as paidingCashAmount," \
                              " sum(case when info.cash_examine_state=3 then info.cash_node_amount else 0 end) as paidCashmount," \
                              " sum(case when info.cash_examine_state=1 then info.card_node_amount else 0 end) as pendingCardAmount," \
                              " sum(case when info.cash_examine_state=2 then info.card_node_amount else 0 end) as paidingCardAmount," \
                              " sum(case when info.cash_examine_state=3 then info.card_node_amount else 0 end) as paidCardmount" \
                              " from brokerage_expendsubject_allocation_info info" \
                              " where info.applyer_id = 387889 and info.allot_type =5 group by subject_id ) sub"
    # 现金奖励
    noPaidCashAmount = "select IFNULL(sum(bai.node_amount),0)" \
                       " from brokerage_award_info bai LEFT JOIN brokerage_award_info_status bais ON bais.brokerage_award_info_id = bai.id" \
                       " LEFT JOIN easylife_order eo ON eo.id = bai.order_id where bai.brokerage_type_id = 1024" \
                       "  AND bai.broker_id = 387889 AND bai.award_type = 2" \
                       " AND (eo.brokerage_confirm_status = 2 or bai.is_mattress = 2) AND (bais.examine_state is null OR bais.examine_state = 1)"

    # 购物卡奖励
    noPaidShoppingCardAmount = "select IFNULL(sum(bai.node_amount),0)" \
                               " from brokerage_award_info bai" \
                               " LEFT JOIN brokerage_award_info_status bais ON bais.brokerage_award_info_id = bai.id" \
                               " LEFT JOIN easylife_order eo ON eo.id = bai.order_id" \
                               " where bai.brokerage_type_id = 1024" \
                               " AND bai.broker_id = 387889" \
                               " AND bai.award_type = 1  AND (eo.brokerage_confirm_status = 2 or bai.is_mattress = 2) AND (bais.examine_state is null OR bais.examine_state = 1)"
    DBSCLISTDESC = ['noPaidCommissionAmountOld', 'noPaidCommissionAmountNew', 'subjectPendingAmountOld',
                    'subjectPendingAmountNew', 'paidingCommissionAmountOld', 'paidingCommissionAmountNew',
                    'paidCommissionAmountOld', 'getKingdeePaidCashAmount', 'getConfirmPaidShoppingCardAmount',
                    'newPaidCommissionAmount', 'noPaidCashAmount', 'noPaidShoppingCardAmount']
    DBSCLIST = [noPaidCommissionAmountOld, noPaidCommissionAmountNew, subjectPendingAmountOld, subjectPendingAmountNew,
                paidingCommissionAmountOld, paidingCommissionAmountNew, paidCommissionAmountOld,
                getKingdeePaidCashAmount, getConfirmPaidShoppingCardAmount, newPaidCommissionAmount, noPaidCashAmount,
                noPaidShoppingCardAmount]
    DBSCLISTDESCRESULT = ["noPaidCommissionAmount", "subjectPendingAmount", "paidingCommissionAmount",
                          "paidCommissionAmount", "noPaidCashAmount", "noPaidShoppingCardAmount"]
# mangodb
# 牵头人佣金测试
class SQLInitiator:
    # 未提交佣金(NotAmountgetNotAmount1+NotAmountgetStatisticsAmountnode_amount+NotAmountgetStatisticsAmountcash_node_amount+NotAmountgetStatisticsAmountcard_node_amount)
    NotAmountgetNotAmount1 = "SELECT IFNULL(SUM(IFNULL(front_node_company_amount, 0.00) + IFNULL(after_node_company_amount, 0.00)),0) " \
                             "FROM brokerage_info WHERE company_id = 933 and examine_state = 1 AND brokerage_type_id = 1024 AND" \
                             "( CASE WHEN node_type = 1 THEN 1 = 1 WHEN node_type = 2 THEN is_completed = 1 END )"
    NotAmountgetStatisticsAmountnode_amount = "select sum(node_amount) from brokerage_expendsubject_allocation_info where allot_type = 2 and org_id = 933 and cash_examine_state=1 and card_examine_state=1"
    NotAmountgetStatisticsAmountcash_node_amount = "select sum(cash_node_amount) from brokerage_expendsubject_allocation_info where allot_type = 2 and org_id = 933 and cash_examine_state=1 and card_examine_state!=1"
    NotAmountgetStatisticsAmountcard_node_amount = "select sum(card_node_amount) from brokerage_expendsubject_allocation_info where allot_type = 2 and org_id = 933 and  card_examine_state=1 and cash_examine_state!=1 "
    # 应结用 shouldAmount(shouldAmountgetShouldAmount+shouldAmountnode_amount)
    shouldAmountgetShouldAmount = "SELECT IFNULL(SUM(IFNULL(front_node_company_amount, 0.00) + IFNULL(after_node_company_amount, 0.00)),0)" \
                                  " FROM brokerage_info WHERE company_id = 933 AND brokerage_type_id = 1024 AND (" \
                                  "  CASE WHEN node_type = 1 THEN 1 = 1 WHEN node_type = 2 THEN is_completed = 1 END )"
    shouldAmountnode_amount = "select sum(node_amount) from brokerage_expendsubject_allocation_info where allot_type = 2 and org_id = 933 and cash_examine_state=1 and card_examine_state=1"
    # 审核中auditAmount(auditAmountgetInAuditAmount+auditAmountnode_amount+auditAmountcash_node_amount+auditAmountcard_node_amount)
    auditAmountgetInAuditAmount = "SELECT IFNULL(SUM(IFNULL(front_node_company_amount, 0.00) + IFNULL(after_node_company_amount, 0.00)),0)" \
                                  " FROM brokerage_info WHERE company_id = 933 and examine_state = 2 and brokerage_type_id = 1024 AND" \
                                  "( CASE WHEN node_type = 1 THEN 1 = 1  WHEN node_type = 2 THEN is_completed = 1 END )"
    auditAmountnode_amount = "select sum(node_amount) from brokerage_expendsubject_allocation_info where cash_examine_state=2 and card_examine_state=2 and allot_type = 2 and org_id = 933"
    auditAmountcash_node_amount = "select sum(cash_node_amount) from brokerage_expendsubject_allocation_info where cash_examine_state=2 and card_examine_state!=2 and allot_type = 2 and org_id = 933"
    auditAmountcard_node_amount = "select sum(card_node_amount) from brokerage_expendsubject_allocation_info where cash_examine_state!=2 and card_examine_state=2 and allot_type = 2 and org_id = 933"
    # 已到账 alreadyAmount(alreadyAmountgetArrivalAmount+alreadyAmountnode_amount+alreadyAmountcash_node_amount+alreadyAmountcard_node_amount)
    alreadyAmountgetArrivalAmount = "SELECT IFNULL(SUM(IFNULL(front_node_company_amount, 0.00) + IFNULL(after_node_company_amount, 0.00)),0)" \
                                    " FROM brokerage_info WHERE company_id = 933 and examine_state = 3  and brokerage_type_id = 1024 AND" \
                                    " ( CASE WHEN node_type = 1 THEN 1 = 1 WHEN node_type = 2 THEN is_completed = 1 END )"
    alreadyAmountnode_amount = "select sum(node_amount) from brokerage_expendsubject_allocation_info where cash_examine_state=3 and card_examine_state=3 and allot_type = 3 and org_id = 933"
    alreadyAmountcash_node_amount = "select sum(cash_node_amount) from brokerage_expendsubject_allocation_info where cash_examine_state=3 and card_examine_state!=3 and allot_type = 3 and org_id = 933"
    alreadyAmountcard_node_amount = "select sum(card_node_amount) from brokerage_expendsubject_allocation_info where cash_examine_state!=3 and card_examine_state=3 and allot_type = 3 and org_id = 933"
    # 预计结佣estimateAmount(estimateAmountgetEstimateAmount+estimateAmountgetEstimateAount)
    estimateAmountgetEstimateAmount = "select IFNULL(SUM(estimateAmount),0) FROM ( select a.order_id as orderId,a.brokerage_node as brokerageNode," \
                                      " (IFNULL(a.front_amount_company,0.00)+IFNULL(a.after_amount_company,0.00)) as estimateAmount" \
                                      " from (select order_id,brokerage_node,front_amount_company,after_amount_company, ( CASE node_type WHEN 1 THEN completed_time" \
                                      " WHEN 2 THEN ( CASE is_completed WHEN 1 THEN completed_time ELSE creat_time END )" \
                                      "  ELSE creat_time end ) AS sortTime from brokerage_info where  company_id = 933 and is_return != 1 and brokerage_type_id = 1024" \
                                      " and ( CASE WHEN node_type = 1 THEN 1 = 1  WHEN node_type = 2 THEN is_completed = 1 END )" \
                                      " order by sortTime desc)  a group by a.order_id)b"
    estimateAmountgetEstimateAount = "select IFNULL(SUM(node_amount),0) from brokerage_expenditure_pre_allot_info where latest_flag = 1 and allot_type = 2 and org_id = 933"
    # 待结佣 surplusAmount(surplusAmountgetEstimateAmount+surplusAmountgetShouldAmount+surplusAmountgetEstimateAount+surplusAmountnode_amount)
    surplusAmountgetEstimateAmount = "select IFNULL(SUM(estimateAmount),0) FROM ( select a.order_id as orderId,a.brokerage_node as brokerageNode," \
                                     "(IFNULL(a.front_amount_company,0.00)+IFNULL(a.after_amount_company,0.00)) as estimateAmount" \
                                     " from (select order_id,brokerage_node,front_amount_company,after_amount_company," \
                                     " ( CASE node_type WHEN 1 THEN completed_time WHEN 2 THEN" \
                                     " ( CASE is_completed WHEN 1 THEN  completed_time ELSE creat_time  END" \
                                     " ) ELSE creat_time end ) AS sortTime from brokerage_info where  company_id = 933 and is_return != 1" \
                                     " and brokerage_type_id = 1024 and " \
                                     "( CASE WHEN node_type = 1 THEN 1 = 1 WHEN node_type = 2 THEN is_completed = 1 END ) order by sortTime desc)  a group by a.order_id)b"
    surplusAmountgetShouldAmount = "SELECT IFNULL(SUM(IFNULL(front_node_company_amount, 0.00) + IFNULL(after_node_company_amount, 0.00)),0)" \
                                   " FROM brokerage_info WHERE company_id = 933 AND brokerage_type_id = 1024 AND " \
                                   "( CASE WHEN node_type = 1 THEN 1 = 1 WHEN node_type = 2 THEN is_completed = 1 END )"
    surplusAmountgetEstimateAount = "select IFNULL(SUM(node_amount),0) from brokerage_expenditure_pre_allot_info" \
                                    " where latest_flag = 1 and allot_type = 2 and org_id = 933"
    surplusAmountnode_amount = "select sum(node_amount) from brokerage_expendsubject_allocation_info where  allot_type = 2 and org_id = 933"
    DBSCLISTDESC = ['NotAmountgetNotAmount1', 'NotAmountgetStatisticsAmountnode_amount',
                    'NotAmountgetStatisticsAmountcash_node_amount', 'NotAmountgetStatisticsAmountcard_node_amount',
                    'shouldAmountgetShouldAmount', 'shouldAmountnode_amount', 'auditAmountgetInAuditAmount',
                    'auditAmountnode_amount', 'auditAmountcash_node_amount', 'auditAmountcard_node_amount',
                    'alreadyAmountgetArrivalAmount',
                    'alreadyAmountnode_amount', 'alreadyAmountcash_node_amount', 'alreadyAmountcard_node_amount',
                    'estimateAmountgetEstimateAmount', 'estimateAmountgetEstimateAount',
                    'surplusAmountgetEstimateAmount', 'surplusAmountgetShouldAmount', 'surplusAmountgetEstimateAount',
                    'surplusAmountnode_amount']
    DBSCLIST = [NotAmountgetNotAmount1, NotAmountgetStatisticsAmountnode_amount,
                NotAmountgetStatisticsAmountcash_node_amount, NotAmountgetStatisticsAmountcard_node_amount,
                shouldAmountgetShouldAmount, shouldAmountnode_amount, auditAmountgetInAuditAmount,
                auditAmountnode_amount, auditAmountcash_node_amount, auditAmountcard_node_amount,
                alreadyAmountgetArrivalAmount,
                alreadyAmountnode_amount, alreadyAmountcash_node_amount, alreadyAmountcard_node_amount,
                estimateAmountgetEstimateAmount, estimateAmountgetEstimateAount,
                surplusAmountgetEstimateAmount, surplusAmountgetShouldAmount, surplusAmountgetEstimateAount,
                surplusAmountnode_amount]
    DBSCLISTDESCRESULT = ["notAmount", "shouldAmount", "auditAmount",
                          "alreadyAmount", "estimateAmount", "surplusAmount"]
class ConfigCommon:
    excelFile=DealExcelTool().getTestFileName()
    sheetName='testCase'
# 全局存储
class FileDB:
    global OPERATIONCODE
# 读取数据，报告写入
class DataDeal:
    def __init__(self):
        pass
    # 读取用例
    # 业务编号	用例编号	用例名称	定位类型	定位内容	操作	输入内容	执行结果	失败明细
    # operationCode testCode testName locationType locationContent operation typeContent executiveOutcome failDetail
    def excelDataGet(self):
        pass
    # 输入报告resultList 为列表类型（结果，详情）
    def excelDataReport(self,et,resultList,listOperation):
        i = listOperation[2] - listOperation[1]
        resultListIndex=0
        for i in range(listOperation[2]):
            et.writeCellValue(row=i, column=7, rcValue=resultList[resultListIndex][0])
            et.writeCellValue(row=i, column=8, rcValue=resultList[resultListIndex][1])
            resultListIndex += 1
    # 兼容unitest
    def unitestReport(self):
        pass
    # 读取Excel的数据，存放到全局变量OPERATIONCODE
    def getTestOperationCode(self,et):
        operationCode=et.get_col_value(1)
        markOperationCode=self.getCodeDeal(operationCode)
        FileDB.OPERATIONCODE=markOperationCode
    # 获取手柄
    def getExcelHandle(self):
        et = ExcelTool(excelFile=ConfigCommon.excelFile, sheetName=ConfigCommon.sheetName)
        return et
    #根据编号标记获取Excel的用例,返回列表
    def getTestData(self,et,listOperation):
        dataListAll=[]
        i = listOperation[2]-listOperation[1]
        for i in range(listOperation[2]):
            dataList=et.get_row_value(row=i, columnNum=1)
            dataListAll.append(dataList)
        return dataListAll
    # 读取业务编号做标记
    def getCodeDeal(self,listData):
        operationCodeValuesdic = {}
        operationCodeValuesdic["list"] = []
        operationCodeValuesdic["listDetail"] = []
        operationCodeValueBn = False
        operationCodeValue = []
        i_num = 1
        i_num_offSet = 1
        a = listData
        for i in a:
            if i in operationCodeValuesdic["list"]:
                i_num += 1
                # 每个业务编号，用例数量，用例开始位置
                operationCodeValue = [i, i_num, i_num_offSet]
            else:
                if i == 3:
                    logger.info(i_num_offSet)
                    logger.info(len(a))
                    logger.info(operationCodeValue)
                if operationCodeValue != [] and (operationCodeValue[0] in operationCodeValuesdic["list"]):
                    operationCodeValuesdic["listDetail"].append(operationCodeValue)
                    operationCodeValue = []
                if operationCodeValue == [] and i_num_offSet == len(a):
                    i_num = 1
                    # 每个业务编号，用例数量，用例开始位置
                    operationCodeValue = [i, i_num, i_num_offSet]
                    operationCodeValuesdic["listDetail"].append(operationCodeValue)
                operationCodeValuesdic["list"].append(i)
                i_num = 1
            i_num_offSet += 1
        return operationCodeValuesdic
    
class Test:
    def __init__(self):
        pass
    
    # 获取Excel的数据
    def readExcelData(self):
        dd = DataDeal()
        et=dd.getExcelHandle()
        dd.getTestOperationCode(et)
        print(FileDB.OPERATIONCODE)
    def getDriver(self):
        driver = GetDriver().chooseDriver()
        return driver
    
    def dbGet(self):
        ms = MysqlSetting(env="t")
        sqlList = DBsqlCommissionTest.DBSCLIST
        data = ms.selectDataSql(sqlList)
        print(data)
    
    def loginApp(self, driver):
        sd = SpiderDemo()
        driver = sd.startApp()
        ta = ToolsAppium()
        driver = sd.loginApp(driver, phone='15301161182')
        return driver
    
    def dbGetMine(self):
        ms = MysqlSetting(env="t")
        sqlList = DBsqlCommissionMine.DBSCLIST
        ms.selectDataSqlMine(sqlList)
    
    def myCommissiondb(self):
        sd = SpiderDemo()
        driver = sd.startApp()
        driver = sd.loginApp(driver, phone='15301161182')
        driver = sd.myCommissionAssert(driver)
        print(driver)
    
    # 我的佣金测试
    def mytest001(self):
        sd = SpiderDemo()
        driver = sd.startApp()
        driver = sd.loginApp(driver, phone='19905021001')
        driver = sd.myCommissionAssert(driver)
        print(driver)
    
    # 牵牛人佣金测试
    def mytest002(self):
        sd = SpiderDemo()
        driver = sd.startApp()
        driver = sd.loginApp(driver, phone='19905021001')
        driver = sd.initiatorCommissionAssert(driver)
        print(driver)
    
    def initiatorTest(self):
        ms = MysqlSetting()
        sqlList = SQLInitiator.DBSCLIST
        data = ms.selectDataSqlInitiator(sqlList)
        print(data)
    
    # 我的佣金测试
    def testMyCommissionAssert(self):
        sd = SpiderDemo()
        driver = sd.startApp()
        driver = sd.loginApp(driver, phone='19905021001')
        data = sd.myCommissionAssert(driver)
        print(data)
    
    # 我的佣金测试
    def testMyCommissionAssert2(self):
        sd = SpiderDemo()
        ut = UiTest()
        driver = sd.startApp(self.getDriver())
        driver = sd.loginApp(driver, phone='19905021001')
        data = ut.testInitiatorCommissionAssert(driver)
        print(data)
    
    # 我的佣金测试
    def testMyCommissionAssert3(self):
        sd = SpiderDemo()
        ut = UiTest()
        driver = sd.startApp(self.getDriver())
        driver = sd.loginApp(driver, phone='19905021001')
        data = ut.testMyCommissionAssert(driver)
        print(data)
    
    # 测试企业认证
    def testMyCommissionAssert4(self):
        sd = SpiderDemo()
        ut = UiTest()
        driver = sd.startApp(self.getDriver())
        driver = sd.loginApp(driver, phone='19905021001')
        data = ut.testCompareEnterpriseCertificationT(driver, xinTax='91350000158142711F', phone='18511225566')
        print(data)
    
    # companyAuthencitionInfomation 查询
    def testMyCommissionAssert5(self):
        sd = SpiderDemo()
        # ut = UiTest()
        driver = sd.startApp(self.getDriver())
        driver = sd.loginApp(driver, phone='18511225564')
        # driver = sd.homeGoMine(driver)
        driver = sd.companyAuthencitionInfomation(driver)
        print(driver)

if __name__ == "__main__":
    T = Test()
    driver = T.readExcelData()
    print(driver)


