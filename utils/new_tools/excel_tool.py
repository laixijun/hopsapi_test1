import os
import re

import openpyxl

from utils.config_tool.ConfigParameter import WebSelenium
from utils.config_tool.file_config_path import ExcelConfig
from utils.logger import Log

logger = Log(logger='excel_tool').get_log()
class ExcelTool:
	def __init__(self,excelFile,sheetName=None):
		self.excelFile=excelFile
		if sheetName != None:
			self.sheetName = sheetName
		else:
			self.sheetName = "Sheet1"
		self.work_book = openpyxl.load_workbook(self.excelFile)
	#获取工作表
	def getSheetValue(self):
		sheet = self.work_book[self.sheetName]
		return sheet

	# row = xl_sheet.max_row 获取行数
	def  getMaxRows(self):
		maxNum=self.getSheetValue().max_row
		return maxNum
	# column = xl_sheet.max_column 获取列数
	def getMaxColumn(self):
		maxColumnNum = self.getSheetValue().max_column
		return maxColumnNum

	# 获取表格的总行数和总列数
	def get_row_clo_num(self):
		rows = self.getSheetValue().max_row
		columns = self.getSheetValue().max_column
		return rows, columns

	# 获取某列的所有值
	def get_col_value(self, column, rowNum=1):
		rows = self.getSheetValue().max_row
		column_data = []
		for i in range(rowNum, rows + 1):
			cell_value = self.getSheetValue().cell(row=i, column=column).value
			if cell_value != None:
				column_data.append(cell_value)
			else:
				break
		return column_data

	# 获取某行所有值
	def get_row_value(self, row, columnNum=1):
		columns = self.getSheetValue().max_column
		row_data = []
		for i in range(columnNum, columns + 1):
			cell_value = self.getSheetValue().cell(row=row, column=i).value
			if cell_value != None:
				row_data.append(cell_value)
			else:
				break
		return row_data

	# 获取

	#读取指定位置的值
	def getCellValue(self,row,column):
		cellValue=self.getSheetValue().cell(row,column).value
		return cellValue

	#获取所有行
	def getAllRowDataToTuple(self):
		rows = self.getSheetValue().rows
		cases=[]
		for row in rows:
			row_cases=[]
			for cell in row:
				row_cases.append(cell.value)
			cases.append(tuple(row_cases))
		return cases

	#向指定位置写入数据
	def writeCellValue(self,row,column,rcValue):
		cellIndex=self.getSheetValue().cell(row,column)
		if not isinstance(rcValue,str):
			rcValue = str(rcValue)
		cellIndex.value=rcValue

	# 向一行写入一条list
	def writeCellRow(self,list,row,column=1):
		column = column
		for i in list:
			self.writeCellValue(row=row,column=column,rcValue=i)
			column += 1

	#将工作簿保存
	def saveWorkbook(self,pathFile):
		self.work_book.save(pathFile)
		self.work_book.close()

class DealExcelTool:
	def __init__(self):
		pass
	# 读取数据的业务ID数据存放入列表
	def readOpretion(self,idOp,startPostition,endPosition):
		pass

	# 获取用例文件的全路径
	def getTestFileName(self):
		testFileName=self.getFilePath()+'/'+ ExcelConfig.TESTCASEALLFile
		testFileName = testFileName.replace('\\','/')
		return testFileName
	# 获取报告文件的全路径
	def getReportFileName(self):
		reportFileName=self.getFileReport() + '/' + ExcelConfig.REPORTPATHFILECURRENT
		reportFileName = reportFileName.replace('\\','/')
		return reportFileName
	# 获取报告文件的全路径复制前
	def getReportFilePreName(self):
		reportFileName=self.getFileReport() + '/' + ExcelConfig.REPORTPATHFILE
		reportFileName = reportFileName.replace('\\','/')
		return reportFileName
	# 获取临时存储文件的全路径
	def getTempFileName(self):
		tempDBFileName=self.getProjectPath() +  ExcelConfig.TEMPDBFILEPATH
		tempDBFileName = tempDBFileName.replace('\\','/')
		return tempDBFileName
	# 获取mac的Firefox driver的全路径
	def getMacFirefoxDriver(self):
		macFirefoxDriver=self.getProjectPath() + ExcelConfig.CONFIGTOOLPATH + WebSelenium.MACFIREFOXDRIVER
		macFirefoxDriver = macFirefoxDriver.replace('\\','/')
		return macFirefoxDriver
	# 获取windows的Firefox driver的全路径
	def getWindowsFirefoxDriver(self):
		windowsFirefoxDriver = self.getProjectPath() + ExcelConfig.CONFIGTOOLPATH + WebSelenium.WINDOWSFIREFOXDRIVER
		windowsFirefoxDriver = windowsFirefoxDriver.replace('\\', '/')
		return windowsFirefoxDriver
	# 获取用例文件的路径
	def getFilePath(self):
		testCaseFilePath = self.getProjectPath()+ ExcelConfig.TESTCASEALL
		return testCaseFilePath
	# 获取报告文件的路径
	def getFileReport(self):
		reportFilePath=self.getProjectPath()+ ExcelConfig.REPORTPATH
		return reportFilePath
	# 获取项目路径
	def getProjectPath(self):
		projectPath=ExcelConfig.PROJECTPATH
		root_path = os.path.abspath(os.path.dirname(__file__))
		reform = re.compile("/(.*?)/utils/new_tools",re.S)
		projectPath = reform.findall(root_path)[0]
		root_path=root_path.split(projectPath)[0]
		# proPath = root_path[0]
		# logger.info(proPath)
		proPath = root_path + projectPath
		return proPath


if __name__ == '__main__':
	a=DealExcelTool()
	c= a.getProjectPath()
	b=a.getProjectPath()
	# root_path = os.path.abspath(os.path.dirname(__file__))
	print(b)
	print(c)